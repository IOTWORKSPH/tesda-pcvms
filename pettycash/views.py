#PettyCash views.py codes
# ==========================================================
# DJANGO CORE IMPORTS
# ==========================================================
from django.shortcuts import render, get_object_or_404, redirect
from django.http import (
    HttpResponse,
    JsonResponse,
    HttpResponseForbidden,
    QueryDict,
)
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group
from django.views.decorators.http import require_POST
from django.contrib import messages

from django.db import transaction
from django.db.models import Q, Sum, F, Min, Max

from django.utils import timezone
from django.utils.dateparse import parse_date
from django.core.paginator import Paginator
from django.conf import settings


# ==========================================================
# PYTHON STANDARD LIBRARIES
# ==========================================================
import os
import zipfile
from io import BytesIO
from decimal import Decimal, InvalidOperation
from copy import copy


# ==========================================================
# THIRD-PARTY LIBRARIES
# ==========================================================
from openpyxl import load_workbook


# ==========================================================
# LOCAL APP MODELS (PETTY CASH)
# ==========================================================
from pettycash.models import (
    PettyCashVoucher,
    VoucherStatus,
    PCVItem,
    Supplier,
    ReceiptAttachment,
    ExpenseCategory,
    TransactionType,
    LiquidationReview,
    Replenishment,
    ReplenishmentStatus,
    Notification,
)


# ==========================================================
# LOCAL APP FORMS
# ==========================================================
from .forms import (
    CashAdvanceForm,
    RefundForm,
    PCVItemFormSet,
)


# ==========================================================
# RELATED APP MODELS
# ==========================================================
from finance.models import PettyCashFund, LedgerEntry, ReferenceType
from audit.models import AuditLog, AuditAction
from users.models import User


# ==========================================================
# SERVICES (BUSINESS LOGIC LAYER)
# ==========================================================
from pettycash.services.workflow_service import WorkflowService
from pettycash.services.replenishment_service import ReplenishmentService
from pettycash.services.replenishment_workflow_service import ReplenishmentWorkflowService

from finance.services.ledger_service import LedgerService

from pettycash.services.replenishment_builder import build_replenishment_context
from pettycash.services.replenishment_pdf import generate_replenishment_pdf


# ==========================================================
# EXCEL GENERATORS (TEMPLATES)
# ==========================================================
from pettycash.services.excel.pcv_excel import generate_pcv_excel
from pettycash.services.excel.pr_excel import generate_pr_excel
from pettycash.services.excel.iar_excel import generate_iar_excel
from pettycash.services.excel.cnrr_excel import generate_cnrr_excel
from pettycash.services.excel.replenishment_excel import generate_replenishment_excel


def notify(user, voucher, message):
    Notification.objects.create(
        user=user,
        voucher=voucher,
        message=message,
        is_read=False
    )

@login_required
def notification_redirect(request, pk):

    notification = get_object_or_404(
        Notification,
        id=pk,
        user=request.user
    )

    notification.is_read = True
    notification.save()

    return redirect("pettycash:pcv_detail", uuid=notification.voucher.uuid)


@login_required
def mark_all_notifications_read(request):

    if request.method == "POST":
        Notification.objects.filter(
            user=request.user,
            is_read=False
        ).update(is_read=True)

    return redirect(request.META.get("HTTP_REFERER", "/"))


@login_required
def notifications_list(request):

    search_query = request.GET.get("q")

    notifications = Notification.objects.filter(
        user=request.user
    ).select_related("voucher").order_by("-created_at")

    # Optional search
    if search_query:
        notifications = notifications.filter(
            Q(message__icontains=search_query) |
            Q(voucher__pcv_no__icontains=search_query)
        )

    # Pagination (10 per page)
    paginator = Paginator(notifications, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "search_query": search_query,
    }

    return render(
        request,
        "pettycash/notifications_list.html",
        context
    )


@login_required
def supplier_autocomplete(request):
    term = request.GET.get("term", "")
    entity = request.user.entity

    suppliers = Supplier.objects.filter(
        entity=entity,
        name__icontains=term
    )[:10]

    data = [
        {"id": s.id, "label": s.name, "value": s.name}
        for s in suppliers
    ]

    return JsonResponse(data, safe=False)


@login_required
@transaction.atomic
def submit_cash_advance(request, uuid):

    if request.method != "POST":
        return redirect("users:dashboard_staff")

    pcv = get_object_or_404(
        PettyCashVoucher,
        uuid=uuid,
        requester=request.user
    )

    # Only cash advance allowed
    if pcv.transaction_type != TransactionType.CASH_ADVANCE:
        messages.error(request, "Invalid transaction type.")
        return redirect("pettycash:pcv_detail", uuid=pcv.uuid)

    # Only draft allowed
    if pcv.status != VoucherStatus.DRAFT:
        messages.error(request, "Only draft requests can be submitted.")
        return redirect("pettycash:pcv_detail", uuid=pcv.uuid)

    if pcv.amount_requested <= 0:
        messages.error(request, "Amount must be greater than zero.")
        return redirect("pettycash:pcv_detail", uuid=pcv.uuid)

    WorkflowService.submit_for_approval(pcv, request.user)

    # 🔔 Notify all Administrators in same entity
    administrators = User.objects.filter(
        entity=pcv.entity,
        groups__name="Administrator",
        is_active=True
    ).distinct()

    for admin in administrators:
        notify(
            admin,
            pcv,
            f"New cash advance request {pcv.pcv_no or '(Pending No.)'} submitted for approval."
        )

    messages.success(request, "Cash advance submitted for approval.")

    return redirect("pettycash:pcv_detail", uuid=pcv.uuid)



@login_required
@transaction.atomic
def submit_refund(request, uuid):

    if request.method != "POST":
        return redirect("users:dashboard_staff")

    pcv = get_object_or_404(
        PettyCashVoucher.objects.select_for_update(),
        uuid=uuid,
        requester=request.user
    )

    # 🔐 Only reimbursement draft allowed
    if pcv.transaction_type != TransactionType.REIMBURSEMENT:
        messages.error(request, "Invalid transaction type.")
        return redirect("pettycash:pcv_detail", uuid=pcv.uuid)

    if pcv.status != VoucherStatus.DRAFT:
        messages.error(request, "Only draft refunds can be submitted.")
        return redirect("pettycash:pcv_detail", uuid=pcv.uuid)

    if not pcv.items.exists():
        messages.error(request, "Cannot submit without items.")
        return redirect("pettycash:pcv_detail", uuid=pcv.uuid)

    # if not pcv.receipts.exists():
    #     messages.error(request, "Receipt attachment is required.")
    #     return redirect("pettycash:pcv_detail", uuid=pcv.uuid)

    WorkflowService.submit_for_approval(pcv, request.user)

    # 🔔 Notify all Administrators in same entity
    administrators = User.objects.filter(
        entity=pcv.entity,
        groups__name="Administrator",
        is_active=True
    ).distinct()

    for admin in administrators:
        notify(
            admin,
            pcv,
            f"New refund request {pcv.pcv_no} submitted for approval."
        )

    messages.success(request, "Refund submitted for approval.")

    return redirect("pettycash:pcv_detail", uuid=pcv.uuid)

@login_required
def create_cash_advance(request):

    if request.method == "POST":
        form = CashAdvanceForm(request.POST, user=request.user)

        if form.is_valid():
            pcv = form.save(commit=False)

            pcv.entity = request.user.entity
            pcv.requester = request.user
            pcv.transaction_type = "CASH_ADVANCE"
            pcv.status = "DRAFT"

            pcv.save()

            messages.success(
                request,
                "Cash advance request saved successfully."
            )

            return redirect("users:dashboard_staff")

    else:
        form = CashAdvanceForm(user=request.user)

    return render(
        request,
        "pettycash/create_cash_advance.html",
        {"form": form}
    )


@login_required
@transaction.atomic
def liquidate_cash_advance(request, uuid):

    pcv = get_object_or_404(
        PettyCashVoucher.objects.select_for_update(),
        uuid=uuid,
        requester=request.user
    )

    if pcv.status != VoucherStatus.RELEASED:
        messages.error(request, "Cash must be released before liquidation.")
        return redirect("pettycash:pcv_detail", uuid=pcv.uuid)

    if request.method == "POST":

        supplier_name = request.POST.get("supplier")
        or_number = request.POST.get("official_receipt_number")
        purchase_date = request.POST.get("purchase_date")
        receipt = request.FILES.get("receipt")

        descriptions = request.POST.getlist("item_description[]")
        quantities = request.POST.getlist("item_qty[]")
        units = request.POST.getlist("item_unit[]")
        costs = request.POST.getlist("item_unit_cost[]")

        # ================= VALIDATION =================

        if not supplier_name:
            messages.error(request, "Supplier name is required.")
            return redirect(request.path)

        if not or_number:
            messages.error(request, "Invoice / OR number is required.")
            return redirect(request.path)

        if not purchase_date:
            messages.error(request, "Purchase date is required.")
            return redirect(request.path)

        # if not receipt and not pcv.receipts.exists():
        #     messages.error(request, "Receipt upload is required.")
        #     return redirect(request.path)

        # ================= PROCESS ITEMS =================

        pcv.items.all().delete()

        total = Decimal("0.00")
        valid_items = 0

        for desc, qty, unit, cost in zip(descriptions, quantities, units, costs):

            if desc.strip():

                try:
                    qty = Decimal(qty)
                    cost = Decimal(cost)
                except:
                    continue

                if qty <= 0 or cost <= 0:
                    continue

                PCVItem.objects.create(
                    voucher=pcv,
                    description=desc.strip(),
                    quantity=qty,
                    unit=unit.strip(),
                    unit_cost=cost
                )

                total += qty * cost
                valid_items += 1

        if valid_items == 0:
            messages.error(request, "At least one valid expense item is required.")
            return redirect(request.path)

        # ================= UPDATE VOUCHER =================

        supplier, _ = Supplier.objects.get_or_create(
            name=supplier_name.strip(),
            entity=request.user.entity,
            defaults={"is_active": True}
        )

        pcv.supplier = supplier
        pcv.official_receipt_number = or_number
        pcv.purchase_date = purchase_date
        pcv.amount_liquidated = total

        # Receipt handling
        if receipt:
            pcv.receipts.all().delete()
            ReceiptAttachment.objects.create(
                voucher=pcv,
                file=receipt,
                uploaded_by=request.user
            )

        pcv.save()

        WorkflowService.liquidate(pcv, request.user)

        messages.success(request, "Liquidation submitted for custodian review.")
        return redirect("pettycash:pcv_detail", uuid=pcv.uuid)

    return render(request, "pettycash/liquidate_cash_advance.html", {
        "pcv": pcv
    })



# @login_required
# @transaction.atomic
# def create_reimbursement(request):

#     entity = request.user.entity

#     categories = ExpenseCategory.objects.filter(
#         entity=entity,
#         is_active=True
#     )

#     funds = PettyCashFund.objects.filter(
#         entity=entity,
#         is_active=True
        
#     )

#     default_fund = funds.first()

#     if request.method == "POST":

#         purchase_date = request.POST.get("purchase_date")
#         purpose = request.POST.get("purpose")
#         expense_category_id = request.POST.get("expense_category")
#         fund_id = request.POST.get("fund")
#         supplier_name = request.POST.get("supplier_name")
#         supplier_id = request.POST.get("supplier_id")
#         invoice_no = request.POST.get("official_receipt_number")
#         receipt_file = request.FILES.get("receipt")

#         descriptions = request.POST.getlist("item_description[]")
#         units = request.POST.getlist("item_unit[]")
#         qtys = request.POST.getlist("item_qty[]")
#         costs = request.POST.getlist("item_unit_cost[]")

#         # =========================
#         # VALIDATION
#         # =========================

#         # if not receipt_file:
#         #     messages.error(request, "Receipt upload is required.")
#         #     return redirect(request.path)

#         if not purchase_date or not purpose:
#             messages.error(request, "Purchase date and purpose are required.")
#             return redirect(request.path)

#         valid_items = []

#         for i in range(len(descriptions)):
#             if descriptions[i] and qtys[i] and costs[i]:
#                 if float(qtys[i]) > 0:
#                     valid_items.append(i)

#         if len(valid_items) == 0:
#             messages.error(request, "At least one valid item is required.")
#             return redirect(request.path)

#         # =========================
#         # SUPPLIER HANDLING
#         # =========================

#         if supplier_id:
#             supplier = Supplier.objects.get(id=supplier_id)
#         else:
#             supplier, created = Supplier.objects.get_or_create(
#                 entity=entity,
#                 name=supplier_name
#             )

#         # =========================
#         # CREATE VOUCHER
#         # =========================

#         pcv = PettyCashVoucher.objects.create(
#             entity=entity,
#             fund_id=fund_id,
#             transaction_type="REIMBURSEMENT",
#             requester=request.user,
#             purpose=purpose,
#             expense_category_id=expense_category_id,
#             purchase_date=purchase_date,
#             supplier=supplier,
#             official_receipt_number=invoice_no,
#             status="DRAFT",
#             amount_requested=Decimal("0.00"),
#             amount_liquidated=Decimal("0.00"),
#         )

#         total_amount = Decimal("0.00")

#         for i in valid_items:
#             qty = Decimal(qtys[i])
#             cost = Decimal(costs[i])
#             line_total = qty * cost

#             PCVItem.objects.create(
#                 voucher=pcv,
#                 description=descriptions[i],
#                 unit=units[i],
#                 quantity=qty,
#                 unit_cost=cost
#             )

#             total_amount += line_total

#         pcv.amount_requested = total_amount
#         pcv.amount_liquidated = total_amount

#         pcv.save()

#         ReceiptAttachment.objects.create(
#             voucher=pcv,
#             file=receipt_file,
#             uploaded_by=request.user
#         )

#         messages.success(request, "Refund request saved successfully.")
#         return redirect("users:dashboard_staff")

#     return render(
#         request,
#         "pettycash/create_reimbursement.html",
#         {
#             "categories": categories,
#             "funds": funds,
#         }
#     )

@login_required
@transaction.atomic
def create_reimbursement(request):

    entity = request.user.entity

    categories = ExpenseCategory.objects.filter(
        entity=entity,
        is_active=True
    )

    funds = PettyCashFund.objects.filter(
        entity=entity,
        is_active=True
    )

    if request.method == "POST":

        purchase_date = request.POST.get("purchase_date")
        purpose = request.POST.get("purpose")
        expense_category_id = request.POST.get("expense_category")
        fund_id = request.POST.get("fund")
        supplier_name = request.POST.get("supplier_name")
        supplier_id = request.POST.get("supplier_id")
        invoice_no = request.POST.get("official_receipt_number")
        receipt_file = request.FILES.get("receipt")

        # ✅ NEW: CNRR FLAG
        has_cnrr = request.POST.get("has_cnrr") == "on"

        descriptions = request.POST.getlist("item_description[]")
        units = request.POST.getlist("item_unit[]")
        qtys = request.POST.getlist("item_qty[]")
        costs = request.POST.getlist("item_unit_cost[]")

        # =========================
        # VALIDATION
        # =========================

        if not purchase_date or not purpose:
            messages.error(request, "Purchase date and purpose are required.")
            return redirect(request.path)

        valid_items = []

        for i in range(len(descriptions)):
            if descriptions[i] and qtys[i] and costs[i]:
                if float(qtys[i]) > 0:
                    valid_items.append(i)

        if len(valid_items) == 0:
            messages.error(request, "At least one valid item is required.")
            return redirect(request.path)

        # =========================
        # COMPUTE TOTAL (FOR CNRR VALIDATION)
        # =========================
        total_amount = Decimal("0.00")

        for i in valid_items:
            qty = Decimal(qtys[i])
            cost = Decimal(costs[i])
            total_amount += qty * cost

        # =========================
        # CNRR RULES
        # =========================
        if has_cnrr:
            if total_amount > Decimal("300"):
                messages.error(request, "CNRR is only allowed for amount ≤ ₱300.")
                return redirect(request.path)

            # Clear OR only (DO NOT TOUCH FILE)
            invoice_no = ""

        # else:
        #     # Optional: enforce receipt if NOT CNRR
        #     if not receipt_file:
        #         messages.error(request, "Receipt is required unless using CNRR.")
        #         return redirect(request.path)

        # =========================
        # SUPPLIER
        # =========================
        if supplier_id:
            supplier = Supplier.objects.get(id=supplier_id)
        else:
            supplier, _ = Supplier.objects.get_or_create(
                entity=entity,
                name=supplier_name
            )

        # =========================
        # CREATE VOUCHER
        # =========================
        pcv = PettyCashVoucher.objects.create(
            entity=entity,
            fund_id=fund_id,
            transaction_type="REIMBURSEMENT",
            requester=request.user,
            purpose=purpose,
            expense_category_id=expense_category_id,
            purchase_date=purchase_date,
            supplier=supplier,
            official_receipt_number=invoice_no,
            status="DRAFT",
            amount_requested=Decimal("0.00"),
            amount_liquidated=Decimal("0.00"),

            # ✅ NEW FIELD
            has_cnrr=has_cnrr,
        )

        # =========================
        # SAVE ITEMS
        # =========================
        for i in valid_items:
            qty = Decimal(qtys[i])
            cost = Decimal(costs[i])

            PCVItem.objects.create(
                voucher=pcv,
                description=descriptions[i],
                unit=units[i],
                quantity=qty,
                unit_cost=cost
            )

        pcv.amount_requested = total_amount
        pcv.amount_liquidated = total_amount
        pcv.save()

        # =========================
        # RECEIPT (ONLY IF NOT CNRR)
        # =========================
        if receipt_file:
            ReceiptAttachment.objects.create(
                voucher=pcv,
                file=receipt_file,
                uploaded_by=request.user
            )

        messages.success(request, "Refund request saved successfully.")
        return redirect("users:dashboard_staff")

    return render(
        request,
        "pettycash/create_reimbursement.html",
        {
            "categories": categories,
            "funds": funds,
        }
    )


@login_required
def print_cnrr(request, uuid):

    voucher = get_object_or_404(
        PettyCashVoucher,
        uuid=uuid,
        entity=request.user.entity
    )

    if not voucher.has_cnrr:
        return render(request, "403.html", status=403)

    administrator = User.objects.filter(
        entity=voucher.entity,
        groups__name="Administrator",
        is_active=True
    ).first()

    wb = generate_cnrr_excel(voucher, administrator)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"CNRR_{voucher.pcv_no or voucher.uuid}.xlsx"

    return HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'inline; filename="{filename}"'
        }
    )


@login_required
def pcv_detail(request, uuid):

    pcv = get_object_or_404(
        PettyCashVoucher,
        uuid=uuid,
        entity=request.user.entity
    )

    approval_logs = AuditLog.objects.filter(
        model_name="PettyCashVoucher",
        object_id=str(pcv.id)
    ).order_by("-created_at")

    return render(
        request,
        "pettycash/pcv_detail.html",
        {
            "pcv": pcv,
            "approval_logs": approval_logs
         }

    )



@login_required
@transaction.atomic
def edit_pcv(request, uuid):

    pcv = get_object_or_404(
        PettyCashVoucher,
        uuid=uuid,
        requester=request.user
    )

    # 🔒 Only DRAFT editable
    if pcv.status != VoucherStatus.DRAFT:
        messages.error(request, "Only draft requests can be edited.")
        return redirect("pettycash:pcv_detail", uuid=pcv.uuid)

    # Determine form
    if pcv.transaction_type == "CASH_ADVANCE":
        FormClass = CashAdvanceForm
    else:
        FormClass = RefundForm

    # ==============================
    # POST
    # ==============================
    if request.method == "POST":

        form = FormClass(
            request.POST,
            instance=pcv,
            user=request.user
        )

        if form.is_valid():

            pcv = form.save(commit=False)

            # =========================================
            # SUPPLIER HANDLING (REFUND ONLY)
            # =========================================
            if pcv.transaction_type == "REIMBURSEMENT":

                supplier_name = form.cleaned_data.get("supplier_name")

                supplier, _ = Supplier.objects.get_or_create(
                    name=supplier_name.strip(),
                    entity=request.user.entity,
                    defaults={"is_active": True}
                )

                pcv.supplier = supplier

            pcv.save()

            # =========================================
            # PROCESS ITEMS (REFUND ONLY)
            # =========================================
            if pcv.transaction_type == "REIMBURSEMENT":

                descriptions = request.POST.getlist("item_description[]")
                quantities = request.POST.getlist("item_qty[]")
                units = request.POST.getlist("item_unit[]")
                costs = request.POST.getlist("item_unit_cost[]")

                # Delete all old items
                pcv.items.all().delete()

                total_amount = Decimal("0.00")

                for desc, qty, unit, cost in zip(descriptions, quantities, units, costs):

                    if desc.strip():

                        try:
                            qty = Decimal(qty)
                            cost = Decimal(cost)
                        except:
                            continue

                        line_total = qty * cost
                        total_amount += line_total

                        PCVItem.objects.create(
                            voucher=pcv,
                            description=desc.strip(),
                            quantity=qty,
                            unit=unit.strip(),
                            unit_cost=cost
                        )

                # Update requested amount automatically for refund
                pcv.amount_requested = total_amount
                pcv.save()

            # =========================================
            # HANDLE RECEIPT REPLACEMENT
            # =========================================
            new_receipt = request.FILES.get("receipt")

            if new_receipt:

                for old in pcv.receipts.all():
                    if old.file:
                        old.file.delete(save=False)
                    old.delete()

                ReceiptAttachment.objects.create(
                    voucher=pcv,
                    file=new_receipt,
                    uploaded_by=request.user
                )

            messages.success(request, "Draft updated successfully.")
            return redirect("pettycash:pcv_detail", uuid=pcv.uuid)

        else:
            messages.error(request, "Please fix the errors below.")

    # ==============================
    # GET
    # ==============================
    else:
        form = FormClass(
            instance=pcv,
            user=request.user
        )

    return render(
        request,
        "pettycash/edit_pcv.html",
        {
            "pcv": pcv,
            "form": form,
        }
    )


@login_required
@transaction.atomic
def approve_voucher(request, uuid):

    # =========================================
    # 🔐 ROLE CHECK
    # =========================================
    if not request.user.has_role("Administrator") and not request.user.is_system_admin:
        return render(request, "403.html", status=403)

    # =========================================
    # 🔒 LOCK VOUCHER (CONCURRENCY SAFE)
    # =========================================
    pcv = get_object_or_404(
        PettyCashVoucher.objects.select_for_update(),
        uuid=uuid,
        entity=request.user.entity
    )

    # =========================================
    # BUSINESS VALIDATION
    # =========================================
    if pcv.status != VoucherStatus.FOR_APPROVAL:
        messages.error(request, "Voucher not eligible for approval.")
        return redirect("pettycash:pcv_detail", uuid=uuid)

    if pcv.amount_requested > pcv.fund.current_balance:
        messages.error(request, "Insufficient fund balance.")
        return redirect("pettycash:pcv_detail", uuid=uuid)

    # =========================================
    # APPROVAL VIA WORKFLOW SERVICE
    # =========================================
    try:
        WorkflowService.approve(pcv, request.user)

        # 🔔 Notify Requester
        notify(
            pcv.requester,
            pcv,
            f"Your request {pcv.pcv_no} has been approved."
        )

        # 🔔 Notify Custodian
        notify(
            pcv.fund.custodian,
            pcv,
            f"{pcv.pcv_no} approved and ready for cash release."
        )

        messages.success(request, "Voucher approved successfully.")

    except PermissionError as e:
        messages.error(request, str(e))

    except ValueError as e:
        messages.error(request, str(e))

    except Exception as e:
        messages.error(request, f"Unexpected error: {str(e)}")

    return redirect("pettycash:pcv_detail", uuid=uuid)



@login_required
@transaction.atomic
def bulk_approve(request):

    if not request.user.has_role("Administrator"):
        return render(request, "403.html", status=403)

    entity = request.user.entity
    selected_ids = request.POST.getlist("selected_vouchers")

    vouchers = PettyCashVoucher.objects.filter(
        id__in=selected_ids,
        entity=entity,
        status=VoucherStatus.FOR_APPROVAL
    )

    # APPROVE ALL
    if request.POST.get("approve_all"):
        vouchers = PettyCashVoucher.objects.filter(
            entity=entity,
            status=VoucherStatus.FOR_APPROVAL
        )

    # ======================
    # BULK REJECT
    # ======================
    if request.POST.get("reject_selected"):

        for voucher in vouchers:
            previous_status = voucher.status
            voucher.status = VoucherStatus.DRAFT
            voucher.save()

            notify(
                voucher.requester,
                voucher,
                f"{voucher.pcv_no} was rejected. Please review."
            )

            AuditLog.objects.create(
                entity=entity,
                user=request.user,
                action=AuditAction.STATUS_CHANGE,
                model_name="PettyCashVoucher",
                object_id=str(voucher.id),
                description=f"{voucher.pcv_no} rejected via bulk.",
                previous_status=previous_status,
                new_status=VoucherStatus.DRAFT
            )

        messages.warning(request, "Selected vouchers rejected.")
        return redirect("users:dashboard_administrator")

    # ======================
    # BULK APPROVE
    # ======================
    approved = 0
    skipped = 0

    for voucher in vouchers:

        if voucher.amount_requested > voucher.fund.current_balance:
            skipped += 1
            continue

        previous_status = voucher.status
        

        try:
            WorkflowService.approve(voucher, request.user)

            notify(voucher.fund.custodian, voucher,
                f"{voucher.pcv_no} approved and ready for cash release.")

            notify(voucher.requester, voucher,
                f"Your request {voucher.pcv_no} has been approved.")

            approved += 1

        except Exception:
            skipped += 1

    if approved:
        messages.success(request, f"{approved} voucher(s) approved.")

    if skipped:
        messages.warning(request, f"{skipped} skipped (insufficient fund).")

    return redirect("users:dashboard_administrator")



@login_required
@transaction.atomic
def reject_voucher(request, uuid):

    if not request.user.has_role("Administrator"):
        return render(request, "403.html", status=403)

    pcv = get_object_or_404(
        PettyCashVoucher.objects.select_for_update(),
        uuid=uuid,
        entity=request.user.entity
    )

    if pcv.status != VoucherStatus.FOR_APPROVAL:
        messages.error(request, "Only vouchers for approval can be rejected.")
        return redirect("pettycash:pcv_detail", uuid=uuid)

    previous_status = pcv.status
    pcv.status = VoucherStatus.DRAFT
    pcv.save()

    # 🔔 Notify Requester
    notify(
        pcv.requester,
        pcv,
        f"{pcv.pcv_no} was rejected. Please review and edit."
    )

    AuditLog.objects.create(
        entity=request.user.entity,
        user=request.user,
        action=AuditAction.STATUS_CHANGE,
        model_name="PettyCashVoucher",
        object_id=str(pcv.id),
        description=f"{pcv.pcv_no} rejected.",
        previous_status=previous_status,
        new_status=VoucherStatus.DRAFT
    )

    messages.warning(request, "Voucher rejected successfully.")
    return redirect("pettycash:pcv_detail", uuid=uuid)



# ==========================================================
# PRINT PCV
# ==========================================================
@login_required
def print_pcv(request, uuid):

    pcv = get_object_or_404(PettyCashVoucher, uuid=uuid)

    entity = pcv.entity

    # Immediate Supervisor = Administrator group
    administrator = User.objects.filter(
        entity=entity,
        groups__name="Administrator",
        is_active=True
    ).first()

    # Petty Cash Custodian = Custodian group
    custodian = User.objects.filter(
        entity=entity,
        groups__name="Custodian",
        is_active=True
    ).first()

    return render(request, "pettycash/print/print_pcv.html", {
        "pcv": pcv,
        "entity": entity,
        "administrator": administrator,
        "custodian": custodian,
    })

# ==========================================================
# PRINT PR
# ==========================================================

@login_required
def print_pr(request, uuid):
    pcv = get_object_or_404(PettyCashVoucher, uuid=uuid)

    # Get Administrator from Django Group
    administrator = User.objects.filter(
        groups__name="Administrator",
        entity=pcv.entity,
        is_active=True
    ).first()

    items = pcv.items.all()

    # Official PR usually has fixed number of rows (e.g., 10)
    total_rows = 10
    blank_count = max(0, total_rows - items.count())

    context = {
        "pcv": pcv,
        "items": items,
        "blank_rows": range(blank_count),
        "entity": pcv.entity,
        "administrator": administrator,
    }

    return render(request, "pettycash/print/print_pr.html", context)


# ==========================================================
# PRINT IINSPECTION AND ACCEPTANCE REPORT
# ==========================================================
@login_required
def print_iar(request, uuid):
    pcv = get_object_or_404(PettyCashVoucher, uuid=uuid)

    # Inspection Committee (Group: Inspection)
    inspectors = User.objects.filter(
        groups__name="Inspection",
        entity=pcv.entity,
        is_active=True
    )

    # Acceptance Officer (Group: Supply Officer)
    acceptance = User.objects.filter(
        groups__name="Supply",
        entity=pcv.entity,
        is_active=True
    ).first()

    items = pcv.items.all()

    total_rows = 8
    blank_count = max(0, total_rows - items.count())

    context = {
        "pcv": pcv,
        "inspectors": inspectors,
        "acceptance": acceptance,
        "blank_rows": range(blank_count),
    }

    return render(request, "pettycash/print/print_iar.html", context)


# ==========================================================
# PRINT ALL DOCUMENTS
# ==========================================================
@login_required
def print_all(request, uuid):

    pcv = get_object_or_404(PettyCashVoucher, uuid=uuid)

    # ==========================================
    # ENTITY
    # ==========================================
    entity = pcv.entity

    # ==========================================
    # ADMINISTRATOR (Immediate Supervisor)
    # ==========================================
    administrator = User.objects.filter(
        groups__name="Administrator",
        entity=entity,
        is_active=True
    ).first()

    # ==========================================
    # CUSTODIAN (Petty Cash Custodian)
    # ==========================================
    custodian = User.objects.filter(
        groups__name="Custodian",
        entity=entity,
        is_active=True
    ).first()

    # ==========================================
    # INSPECTION COMMITTEE
    # ==========================================
    inspectors = User.objects.filter(
        groups__name="Inspection",
        entity=entity,
        is_active=True
    )

    # ==========================================
    # SUPPLY OFFICER (Acceptance)
    # ==========================================
    acceptance = User.objects.filter(
        groups__name="Supply",
        entity=entity,
        is_active=True
    ).first()

    # ==========================================
    # ITEMS
    # ==========================================
    items = pcv.items.all()

    # Blank rows for PR / IAR formatting
    blank_rows = range(max(0, 10 - items.count()))

    context = {
        "pcv": pcv,
        "entity": entity,
        "administrator": administrator,
        "custodian": custodian,
        "inspectors": inspectors,
        "acceptance": acceptance,
        "items": items,
        "blank_rows": blank_rows,
    }

    return render(
        request,
        "pettycash/print/print_all.html",
        context
    )


# ==========================================================
# RELEASE CASH (FOR CASH ADVANCE)
# ==========================================================

@login_required
@require_POST
@transaction.atomic
def release_cash_view(request, pk):

    voucher = get_object_or_404(
        PettyCashVoucher.objects.select_for_update(),
        pk=pk,
        entity=request.user.entity
    )

    try:
        WorkflowService.release_cash(voucher, request.user)

        messages.success(
            request,
            f"Cash successfully released for PCV {voucher.pcv_no}."
        )

    except PermissionError as e:
        messages.error(request, str(e))

    except ValueError as e:
        messages.error(request, str(e))

    except Exception as e:
        messages.error(request, f"Unexpected error: {str(e)}")

    return redirect("users:dashboard_custodian")


# ==========================================================
# POST REIMBURSEMENT
# ==========================================================

@login_required
@require_POST
@transaction.atomic
def post_reimbursement_view(request, pk):

    voucher = get_object_or_404(
        PettyCashVoucher.objects.select_for_update(),
        pk=pk,
        entity=request.user.entity
    )

    try:
        WorkflowService.post_reimbursement(voucher, request.user)

        messages.success(
            request,
            f"Reimbursement successfully paid and posted for PCV {voucher.pcv_no}."
        )

    except PermissionError as e:
        messages.error(request, str(e))

    except ValueError as e:
        messages.error(request, str(e))

    except Exception as e:
        messages.error(request, f"Unexpected error: {str(e)}")

    return redirect("users:dashboard_custodian")


@login_required
def custodian_release_list(request):

    if not request.user.has_role("Custodian"):
        return render(request, "403.html", status=403)

    fund = PettyCashFund.objects.filter(
        custodian=request.user,
        is_active=True
    ).first()

    if not fund:
        return render(request, "users/no_fund.html")

    search = request.GET.get("search", "")
    date_from = request.GET.get("date_from", "")
    date_to = request.GET.get("date_to", "")

    vouchers = PettyCashVoucher.objects.filter(
        fund=fund,
        status=VoucherStatus.APPROVED
    ).select_related("requester")

    # 🔎 SEARCH
    if search:
        vouchers = vouchers.filter(
            Q(pcv_no__icontains=search) |
            Q(requester__first_name__icontains=search) |
            Q(requester__last_name__icontains=search)
        )

    # 📅 DATE RANGE (based on approval date or created_at)
    if date_from:
        vouchers = vouchers.filter(
            created_at__date__gte=parse_date(date_from)
        )

    if date_to:
        vouchers = vouchers.filter(
            created_at__date__lte=parse_date(date_to)
        )

    vouchers = vouchers.order_by("-created_at")

    return render(
        request,
        "pettycash/custodian/release_list.html",
        {
            "fund": fund,
            "vouchers": vouchers,
            "search": search,
            "date_from": date_from,
            "date_to": date_to,
        }
    )



@login_required
def custodian_unliquidated(request):

    if not request.user.has_role("Custodian"):
        return render(request, "403.html", status=403)

    fund = PettyCashFund.objects.filter(
        custodian=request.user,
        is_active=True
    ).first()

    search = request.GET.get("search", "")
    date_from = request.GET.get("date_from", "")
    date_to = request.GET.get("date_to", "")

    vouchers = PettyCashVoucher.objects.filter(
        fund=fund,
        transaction_type=TransactionType.CASH_ADVANCE,
        status=VoucherStatus.RELEASED
    ).select_related("requester")

    if search:
        vouchers = vouchers.filter(
            Q(pcv_no__icontains=search) |
            Q(requester__first_name__icontains=search) |
            Q(requester__last_name__icontains=search)
        )

    if date_from:
        vouchers = vouchers.filter(
            updated_at__date__gte=parse_date(date_from)
        )

    if date_to:
        vouchers = vouchers.filter(
            updated_at__date__lte=parse_date(date_to)
        )

    vouchers = vouchers.order_by("-updated_at")

    return render(
        request,
        "pettycash/custodian/unliquidated.html",
        {
            "fund": fund,
            "vouchers": vouchers,
            "search": search,
            "date_from": date_from,
            "date_to": date_to,
        }
    )



@login_required
def custodian_fund_ledger(request):

    if not request.user.has_role("Custodian"):
        return render(request, "403.html", status=403)

    fund = PettyCashFund.objects.filter(
        custodian=request.user,
        is_active=True
    ).first()

    search = request.GET.get("search", "")
    date_from = request.GET.get("date_from", "")
    date_to = request.GET.get("date_to", "")

    entries = LedgerEntry.objects.filter(fund=fund)

    if search:
        entries = entries.filter(
            Q(reference_no__icontains=search)
        )

    if date_from:
        entries = entries.filter(
            transaction_date__gte=parse_date(date_from)
        )

    if date_to:
        entries = entries.filter(
            transaction_date__lte=parse_date(date_to)
        )

    entries = entries.order_by("-transaction_date")

    return render(
        request,
        "pettycash/custodian/fund_ledger.html",
        {
            "fund": fund,
            "entries": entries,
            "search": search,
            "date_from": date_from,
            "date_to": date_to,
        }
    )




@login_required
@transaction.atomic
def finalize_liquidation(request, uuid):

    if request.method != "POST":
        messages.error(request, "Invalid request method.")
        return redirect("users:dashboard_custodian")

    # 🔐 Custodian only
    if not request.user.groups.filter(name="Custodian").exists():
        messages.error(request, "Unauthorized action.")
        return redirect("users:dashboard_redirect")

    voucher = get_object_or_404(
        PettyCashVoucher.objects.select_for_update(),
        uuid=uuid
    )

    # 🔐 Ensure voucher belongs to custodian fund
    if voucher.fund.custodian != request.user:
        messages.error(request, "You are not assigned to this fund.")
        return redirect("users:dashboard_custodian")

    # 🔎 Must be cash advance
    if voucher.transaction_type != TransactionType.CASH_ADVANCE:
        messages.error(request, "Invalid transaction type.")
        return redirect("users:dashboard_custodian")

    # 🔎 Must already be liquidated by staff
    if voucher.status != VoucherStatus.LIQUIDATED:
        messages.error(request, "Voucher not ready for finalization.")
        return redirect("users:dashboard_custodian")

    try:
        # 🔥 THIS HANDLES ALL FINANCIAL LOGIC
        LedgerService.post_liquidation_adjustment(voucher, request.user)

        messages.success(
            request,
            f"Liquidation finalized and posted for {voucher.pcv_no}."
        )

    except ValueError as e:
        messages.error(request, str(e))

    return redirect("users:dashboard_custodian")

@login_required
@transaction.atomic
def return_liquidation(request, uuid):

    voucher = get_object_or_404(
        PettyCashVoucher.objects.select_for_update(),
        uuid=uuid
    )

    if request.method != "POST":
        return redirect("pettycash:pcv_detail", uuid=uuid)

    if not request.user.has_role("Custodian"):
        return render(request, "403.html", status=403)

    remarks = request.POST.get("remarks")

    if not remarks:
        messages.error(request, "Remarks are required.")
        return redirect("pettycash:pcv_detail", uuid=uuid)

    # 🔥 Save review record
    LiquidationReview.objects.create(
        voucher=voucher,
        reviewed_by=request.user,
        remarks=remarks,
        action="RETURNED"
    )

    # Change status
    voucher.status = VoucherStatus.RELEASED
    voucher.save(update_fields=["status"])

    messages.warning(
        request,
        "Liquidation returned to staff for correction."
    )

    return redirect("pettycash:pcv_detail", uuid=uuid)


@login_required
@require_POST
@transaction.atomic
def delete_pcv(request, uuid):

    pcv = get_object_or_404(
        PettyCashVoucher.objects.select_for_update(),
        uuid=uuid,
        requester=request.user
    )

    # 🔒 Only draft allowed
    if pcv.status != VoucherStatus.DRAFT:
        messages.error(request, "Only draft transactions can be deleted.")
        return redirect("users:dashboard_staff")

    pcv.delete()

    messages.success(request, "Draft transaction deleted successfully.")

    return redirect("users:dashboard_staff")

@login_required
def replenishment_report(request):

    if not request.user.has_role("Custodian"):
        return render(request, "403.html", status=403)

    fund = PettyCashFund.objects.filter(
        custodian=request.user,
        is_active=True
    ).first()

    if not fund:
        return render(request, "users/no_fund.html")

    from_date = request.GET.get("date_from")
    to_date = request.GET.get("date_to")
    search_query = request.GET.get("q", "").strip()

    vouchers = PettyCashVoucher.objects.filter(
        fund=fund,
        status=VoucherStatus.POSTED,
        is_replenished=False,
        replenishment__isnull=True
    )

    if from_date:
        vouchers = vouchers.filter(
            purchase_date__gte=parse_date(from_date)
        )

    if to_date:
        vouchers = vouchers.filter(
            purchase_date__lte=parse_date(to_date)
        )

    if search_query:
        vouchers = vouchers.filter(
            Q(pcv_no__icontains=search_query) |
            Q(purpose__icontains=search_query)
        )

    vouchers = vouchers.order_by("-purchase_date")

    total = vouchers.aggregate(
        total=Sum("amount_requested")
    )["total"] or Decimal("0.00")

    return render(
        request,
        "pettycash/reports/replenishment_report.html",
        {
            "fund": fund,
            "vouchers": vouchers,
            "total": total,
            "date_from": from_date,
            "date_to": to_date,
            "search_query": search_query,
        }
    )


@login_required
def replenishment_generate(request):

    if request.method != "POST":
        return redirect("pettycash:replenishment_report")

    fund = PettyCashFund.objects.filter(
        custodian=request.user,
        is_active=True
    ).first()

    if not fund:
        messages.error(request, "No active fund found.")
        return redirect("pettycash:replenishment_report")

    # Preserve filters (if coming from filtered report)
    date_from = request.POST.get("date_from")
    date_to = request.POST.get("date_to")

    # Base eligible queryset
    vouchers_qs = PettyCashVoucher.objects.filter(
        fund=fund,
        is_posted_to_ledger=True,
        is_replenished=False,
        replenishment__isnull=True
    )

    # Apply date filters if provided
    if date_from:
        vouchers_qs = vouchers_qs.filter(purchase_date__gte=date_from)

    if date_to:
        vouchers_qs = vouchers_qs.filter(purchase_date__lte=date_to)

    # Get selected vouchers
    selected_ids = request.POST.getlist("selected_vouchers")

    # ======================================================
    # ENTERPRISE SELECTION LOGIC
    # ======================================================

    if selected_ids:
        vouchers = vouchers_qs.filter(id__in=selected_ids)
    else:
        # If none selected → include all filtered
        vouchers = vouchers_qs

    if not vouchers.exists():
        messages.error(request, "No eligible transactions found.")
        return redirect("pettycash:replenishment_report")

    total = vouchers.aggregate(
        total=Sum("amount_requested")
    )["total"] or Decimal("0.00")

    return render(
        request,
        "pettycash/reports/replenishment_print.html",
        {
            "fund": fund,
            "vouchers": vouchers.order_by("-release_date"),
            "total": total,
            "date_from": date_from,
            "date_to": date_to,
        }
    )


@login_required
def replenishment_package_pdf(request):

    if request.method != "POST":
        return redirect("pettycash:replenishment_report")

    context = build_replenishment_context(request)

    if not context:
        messages.error(request, "No eligible transactions found.")
        return redirect("pettycash:replenishment_report")

    return generate_replenishment_pdf(request, context)



@login_required
def replenishment_export_excel(request):

    if request.method != "POST":
        return redirect("pettycash:replenishment_report")

    context = build_replenishment_context(request)

    if not context:
        messages.error(request, "No eligible transactions found.")
        return redirect("pettycash:replenishment_report")

    return generate_replenishment_excel(context)



@login_required
@transaction.atomic
def create_replenishment(request):

    if not request.user.has_role("Custodian"):
        return render(request, "403.html", status=403)

    fund = PettyCashFund.objects.filter(
        custodian=request.user,
        is_active=True
    ).first()

    if not fund:
        messages.error(request, "No active fund found.")
        return redirect("users:dashboard_custodian")

    selected_ids = request.POST.getlist("selected_vouchers")

    base_qs = PettyCashVoucher.objects.filter(
        fund=fund,
        status=VoucherStatus.POSTED,
        is_replenished=False,
        replenishment__isnull=True
    )

    if selected_ids:
        vouchers = base_qs.filter(id__in=selected_ids)
    else:
        vouchers = base_qs

    total_expenses = Decimal("0.00")

    for voucher in vouchers:
        total_expenses += voucher.actual_amount

    if request.method == "POST":

        if not vouchers.exists():
            messages.error(request, "No eligible vouchers found.")
            return redirect("pettycash:replenishment_report")

        # ===============================
        # LOCK FUND FOR SAFE SERIES
        # ===============================
        fund = PettyCashFund.objects.select_for_update().get(pk=fund.pk)

        current_year = timezone.now().year

        last_series = (
            Replenishment.objects
            .select_for_update()
            .filter(year=current_year, fund=fund)
            .aggregate(Max("series_number"))
        )["series_number__max"] or 0

        next_series = last_series + 1
        report_number = f"{current_year}-{next_series:04d}"

        # ===============================
        # SNAPSHOT OPENING BALANCE
        # ===============================
        latest_entry = fund.ledger_entries.order_by(
            "-transaction_date", "-id"
        ).first()

        opening_balance = (
            latest_entry.running_balance
            if latest_entry
            else fund.current_balance
        )

        replenishment = Replenishment.objects.create(
            fund=fund,
            year=current_year,
            series_number=next_series,
            report_number=report_number,
            opening_balance=opening_balance,
            total_expenses=total_expenses,
            status=ReplenishmentStatus.DRAFT,
            created_by=request.user
        )

        # Link vouchers (but DO NOT mark replenished yet)
        vouchers.update(replenishment=replenishment)

        messages.success(
            request,
            f"Replenishment {report_number} generated (Draft)."
        )

        return redirect(
            "pettycash:replenishment_detail",
            pk=replenishment.pk
        )

    return render(
        request,
        "pettycash/custodian/create_replenishment.html",
        {
            "fund": fund,
            "vouchers": vouchers,
            "total_expenses": total_expenses
        }
    )



@login_required
def replenishment_list(request):

    if not request.user.has_role("Custodian"):
        return render(request, "403.html", status=403)

    fund = PettyCashFund.objects.filter(
        custodian=request.user,
        is_active=True
    ).first()

    replenishments = Replenishment.objects.filter(
        fund=fund
    ).order_by("-created_at")

    return render(
        request,
        "pettycash/custodian/replenishment_list.html",
        {
            "fund": fund,
            "replenishments": replenishments
        }
    )



@login_required
def replenishment_detail(request, pk):

    replenishment = get_object_or_404(
        Replenishment,
        pk=pk,
        fund__entity=request.user.entity
    )

    # Group checks
    is_admin = request.user.groups.filter(name="Administrator").exists()
    is_custodian = request.user.groups.filter(name="Custodian").exists()

    # Permission check
    if not (is_admin or replenishment.fund.custodian == request.user):
        return render(request, "403.html", status=403)

    vouchers = replenishment.vouchers.all().order_by("pcv_no")

    context = {
        "replenishment": replenishment,
        "vouchers": vouchers,
        "is_admin": is_admin,
        "is_custodian": is_custodian,
    }

    return render(
        request,
        "pettycash/custodian/replenishment_detail.html",
        context
    )


@login_required
@transaction.atomic
def create_initial_fund(request):

    # ==========================================
    # ROLE CHECK
    # ==========================================
    if not request.user.has_role("Custodian"):
        return render(request, "403.html", status=403)

    # ==========================================
    # PREVENT DUPLICATE ACTIVE FUND
    # ==========================================
    existing = PettyCashFund.objects.filter(
        custodian=request.user,
        is_active=True
    ).exists()

    if existing:
        messages.warning(request, "Active fund already exists.")
        return redirect("users:dashboard_custodian")

    # ==========================================
    # HANDLE POST
    # ==========================================
    if request.method == "POST":

        # ---- Safe Decimal Parsing ----
        raw_amount = request.POST.get("fund_amount")

        try:
            fund_amount = Decimal(raw_amount)
            if fund_amount <= 0:
                raise ValueError
        except (InvalidOperation, TypeError, ValueError):
            messages.error(request, "Invalid fund amount.")
            return redirect(request.path)

        funding_mode = request.POST.get("funding_mode")
        check_number = request.POST.get("check_number", "").strip()
        check_date = request.POST.get("check_date")

        # ==========================================
        # VALIDATE CHECK MODE
        # ==========================================
        if funding_mode == "CHECK":
            if not check_number or not check_date:
                messages.error(
                    request,
                    "Check number and check date are required for check funding."
                )
                return redirect(request.path)

        # ==========================================
        # CREATE FUND
        # ==========================================
        fund = PettyCashFund.objects.create(
            entity=request.user.entity,
            fund_cluster_id=request.POST.get("fund_cluster"),
            responsibility_center_id=request.POST.get("responsibility_center"),
            name="Petty Cash Fund",
            custodian=request.user,
            fund_amount=fund_amount,
            current_balance=fund_amount,
            is_active=True
        )

        # ==========================================
        # DETERMINE LEDGER REFERENCE
        # ==========================================
        if funding_mode == "CHECK":
            reference_no = f"CHK-{check_number}"
            description = f"Initial Fund via Check #{check_number}"
        else:
            reference_no = "CASH-OPENING"
            description = "Initial Fund via Cash"

        # ==========================================
        # OPENING LEDGER ENTRY (DEBIT)
        # ==========================================
        LedgerEntry.objects.create(
            fund=fund,
            transaction_date=timezone.now().date(),
            debit=fund_amount,
            credit=Decimal("0.00"),
            running_balance=fund_amount,
            reference_type=ReferenceType.ADJUSTMENT,
            reference_no=reference_no,
            description=description,
            created_by=request.user
        )

        messages.success(request, "Initial fund created successfully.")
        return redirect("users:dashboard_custodian")

    # ==========================================
    # HANDLE GET
    # ==========================================
    from finance.models import FundCluster, ResponsibilityCenter

    clusters = FundCluster.objects.all()
    rccs = ResponsibilityCenter.objects.filter(
        entity=request.user.entity,
        is_active=True
    )

    return render(
        request,
        "pettycash/custodian/create_initial_fund.html",
        {
            "clusters": clusters,
            "rccs": rccs
        }
    )



@login_required
def my_vouchers(request):

    search = request.GET.get("q", "")

    vouchers = (
        PettyCashVoucher.objects
        .filter(requester=request.user)
        .select_related("expense_category")
        .order_by("-created_at")
    )

    if search:
        vouchers = vouchers.filter(
            Q(pcv_no__icontains=search) |
            Q(purpose__icontains=search) |
            Q(expense_category__name__icontains=search)
        )

    paginator = Paginator(vouchers, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "search": search,
    }

    return render(
        request,
        "pettycash/my_vouchers.html",
        context
    )


@login_required
def inspection_pending(request):

    if not request.user.has_role("Inspection"):
        return render(request, "403.html", status=403)

    entity = request.user.entity
    search = request.GET.get("q", "")

    vouchers = PettyCashVoucher.objects.filter(
        entity=entity,
        status=VoucherStatus.APPROVED,
        transaction_type=TransactionType.REIMBURSEMENT,
        is_posted_to_ledger=False
    ).select_related("requester")

    if search:
        vouchers = vouchers.filter(
            Q(pcv_no__icontains=search) |
            Q(requester__first_name__icontains=search) |
            Q(requester__last_name__icontains=search)
        )

    vouchers = vouchers.order_by("-created_at")

    context = {
        "vouchers": vouchers,
        "search": search,
    }

    return render(request, "pettycash/inspection/inspection_pending.html", context)



@login_required
def inspection_all_items(request):

    if not request.user.has_role("Inspection"):
        return render(request, "403.html", status=403)

    entity = request.user.entity
    search = request.GET.get("q", "")

    items = PCVItem.objects.filter(
        voucher__entity=entity,
        voucher__status__in=[
            VoucherStatus.APPROVED,
            VoucherStatus.POSTED,
            VoucherStatus.LIQUIDATED
        ]
    ).select_related("voucher")

    if search:
        items = items.filter(
            Q(description__icontains=search) |
            Q(voucher__pcv_no__icontains=search)
        )

    items = items.order_by("-voucher__created_at")

    context = {
        "items": items,
        "search": search,
    }

    return render(request, "pettycash/inspection/inspection_all_items.html", context)


@login_required
def supply_items(request):

    if not request.user.has_role("Supply"):
        return render(request, "403.html", status=403)

    entity = request.user.entity
    search = request.GET.get("q", "")

    vouchers = PettyCashVoucher.objects.filter(
        entity=entity,
        iar_no__isnull=False
    ).select_related("requester")

    if search:
        vouchers = vouchers.filter(
            Q(pcv_no__icontains=search) |
            Q(iar_no__icontains=search) |
            Q(requester__first_name__icontains=search) |
            Q(requester__last_name__icontains=search)
        )

    vouchers = vouchers.order_by("-created_at")

    context = {
        "vouchers": vouchers,
        "search": search,
    }

    return render(request, "pettycash/supply/supply_items.html", context)


@login_required
def supply_iar_pending(request):

    if not request.user.has_role("Supply"):
        return render(request, "403.html", status=403)

    entity = request.user.entity
    search = request.GET.get("q", "")

    vouchers = PettyCashVoucher.objects.filter(
        entity=entity,
        status=VoucherStatus.APPROVED,
        iar_no__isnull=True
    ).select_related("requester")

    if search:
        vouchers = vouchers.filter(
            Q(pcv_no__icontains=search) |
            Q(requester__first_name__icontains=search) |
            Q(requester__last_name__icontains=search)
        )

    vouchers = vouchers.order_by("-created_at")

    context = {
        "vouchers": vouchers,
        "search": search,
    }

    return render(request, "pettycash/supply/supply_iar_pending.html", context)

@login_required
def generate_iar(request, uuid):

    if not request.user.has_role("Supply"):
        return render(request, "403.html", status=403)

    voucher = get_object_or_404(
        PettyCashVoucher,
        uuid=uuid,
        entity=request.user.entity
    )

    if voucher.iar_no:
        messages.warning(request, "IAR already issued.")
        return redirect("pettycash:supply_iar_pending")

    # Example: IAR-2026-0001 format
    year = timezone.now().year
    count = PettyCashVoucher.objects.filter(
        entity=voucher.entity,
        iar_no__startswith=f"IAR-{year}"
    ).count() + 1

    voucher.iar_no = f"IAR-{year}-{count:04d}"
    voucher.save()

    messages.success(request, "IAR generated successfully.")

    return redirect("pettycash:supply_items")



@login_required
def submit_replenishment(request, pk):

    replenishment = get_object_or_404(
        Replenishment,
        pk=pk,
        fund__custodian=request.user
    )

    try:
        ReplenishmentWorkflowService.submit(replenishment)
        messages.success(request, "Replenishment submitted.")
    except ValueError as e:
        messages.error(request, str(e))

    return redirect("pettycash:replenishment_detail", pk=pk)


@login_required
def approve_replenishment(request, pk):

    if not request.user.has_role("Administrator"):
        return render(request, "403.html", status=403)

    replenishment = get_object_or_404(Replenishment, pk=pk)

    try:
        ReplenishmentWorkflowService.approve(replenishment)
        messages.success(request, "Replenishment approved.")
    except ValueError as e:
        messages.error(request, str(e))

    return redirect("pettycash:replenishment_detail", pk=pk)


@login_required
@require_POST
@transaction.atomic
def release_replenishment(request, pk):

    # 🔐 Custodian only
    if not request.user.has_role("Custodian"):
        return render(request, "403.html", status=403)

    replenishment = get_object_or_404(
        Replenishment.objects.select_for_update(),
        pk=pk,
        fund__entity=request.user.entity
    )

    # 🔐 Ensure assigned fund
    if replenishment.fund.custodian != request.user:
        messages.error(request, "You are not assigned to this fund.")
        return redirect("pettycash:replenishment_detail", pk=pk)

    # 🔎 Must be submitted
    if replenishment.status != ReplenishmentStatus.SUBMITTED_TO_ACCOUNTING:
        messages.error(request, "Replenishment must be submitted first.")
        return redirect("pettycash:replenishment_detail", pk=pk)

    # 📥 Get inputs
    check_number = request.POST.get("check_number", "").strip()
    check_date = request.POST.get("check_date")
    check_amount_raw = request.POST.get("check_amount")

    # 🔎 Validation
    if not check_number:
        messages.error(request, "Check number is required.")
        return redirect("pettycash:replenishment_detail", pk=pk)

    if not check_date:
        messages.error(request, "Check date is required.")
        return redirect("pettycash:replenishment_detail", pk=pk)

    try:
        check_amount = Decimal(check_amount_raw)
    except:
        messages.error(request, "Invalid check amount.")
        return redirect("pettycash:replenishment_detail", pk=pk)

    if check_amount != replenishment.total_expenses:
        messages.error(
            request,
            "Check amount must match total expenses."
        )
        return redirect("pettycash:replenishment_detail", pk=pk)

    try:
        ReplenishmentWorkflowService.release(
            replenishment,
            check_number,
            check_date,
            request.user
        )

        messages.success(
            request,
            f"Replenishment {replenishment.report_number} released successfully."
        )

    except ValueError as e:
        messages.error(request, str(e))

    return redirect("pettycash:replenishment_detail", pk=pk)

@login_required
@require_POST
@transaction.atomic
def delete_replenishment(request, pk):

    replenishment = get_object_or_404(
        Replenishment.objects.select_for_update(),
        pk=pk
    )

    if not request.user.has_role("Custodian"):
        return render(request, "403.html", status=403)

    if replenishment.fund.custodian != request.user:
        return render(request, "403.html", status=403)

    if replenishment.status != ReplenishmentStatus.DRAFT:
        messages.error(
            request,
            "Submitted or released replenishments cannot be deleted."
        )
        return redirect("pettycash:replenishment_detail", pk=pk)

    replenishment.vouchers.update(
        replenishment=None,
        is_replenished=False
    )

    replenishment.delete()

    messages.success(request, "Draft replenishment deleted.")
    return redirect("pettycash:replenishment_list")


def clone_sheet(source_ws, target_wb, title):

    target_ws = target_wb.create_sheet(title=title)

    # Copy cell values + styles
    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = target_ws[cell.coordinate]
            new_cell.value = cell.value

            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.alignment = copy(cell.alignment)

    # ✅ COPY COLUMN WIDTHS
    for col in source_ws.column_dimensions:
        target_ws.column_dimensions[col].width = source_ws.column_dimensions[col].width

    # ✅ COPY ROW HEIGHTS
    for row in source_ws.row_dimensions:
        target_ws.row_dimensions[row].height = source_ws.row_dimensions[row].height

    # ✅ VERY IMPORTANT: COPY MERGED CELLS
    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    return target_ws


@login_required
def replenishment_bundle_excel(request, pk):

    replenishment = get_object_or_404(
        Replenishment,
        pk=pk,
        fund__entity=request.user.entity
    )

    if not (
        request.user.has_role("Custodian")
        or request.user.has_role("Administrator")
    ):
        return render(request, "403.html", status=403)

    vouchers = replenishment.vouchers.all().order_by(
        "purchase_date",
        "pcv_no"
    )

    zip_buffer = BytesIO()

    with zipfile.ZipFile(zip_buffer, "w") as zip_file:

        for voucher in vouchers:

            # ================= SIGNATORIES =================
            administrator = User.objects.filter(
                entity=voucher.entity,
                groups__name="Administrator",
                is_active=True
            ).first()

            custodian = voucher.fund.custodian

            # ================= GENERATE WORKBOOKS =================
            wb_pcv = generate_pcv_excel(
                voucher,
                voucher.entity,
                administrator,
                custodian
            )

            wb_pr = generate_pr_excel(voucher, administrator)
            wb_iar = generate_iar_excel(voucher)

            # Rename PCV sheet
            wb_pcv.active.title = "PCV"

            # Clone PR
            clone_sheet(wb_pr.active, wb_pcv, "PR")

            # Clone IAR
            clone_sheet(wb_iar.active, wb_pcv, "IAR")

            # ✅ CNRR
            if voucher.has_cnrr:
                wb_cnrr = generate_cnrr_excel(voucher, administrator)
                clone_sheet(wb_cnrr.active, wb_pcv, "CNRR")

            

            # ================= SAVE FINAL FILE =================
            buffer = BytesIO()
            wb_pcv.save(buffer)
            buffer.seek(0)

            filename = f"{voucher.pcv_no or voucher.uuid}.xlsx"

            zip_file.writestr(filename, buffer.read())

    zip_buffer.seek(0)

    return HttpResponse(
        zip_buffer,
        content_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="replenishment_{replenishment.report_number}.zip"'
        }
    )

@login_required
def export_voucher_excel(request, uuid):

    voucher = get_object_or_404(
        PettyCashVoucher,
        uuid=uuid,
        entity=request.user.entity
    )

    # ================= SECURITY =================
    if not (
        request.user == voucher.requester
        or request.user.has_role("Administrator")
        or request.user.has_role("Custodian")
    ):
        return render(request, "403.html", status=403)

    # ================= SIGNATORIES =================
    administrator = User.objects.filter(
        entity=voucher.entity,
        groups__name="Administrator",
        is_active=True
    ).first()

    custodian = voucher.fund.custodian if voucher.fund else None

    # ================= GENERATE WORKBOOKS =================
    wb_pcv = generate_pcv_excel(
        voucher,
        voucher.entity,
        administrator,
        custodian
    )

    wb_pr = generate_pr_excel(voucher, administrator)
    wb_iar = generate_iar_excel(voucher)

    # ================= COMBINE INTO ONE WORKBOOK =================
    wb_pcv.active.title = "PCV"

    clone_sheet(wb_pr.active, wb_pcv, "PR")
    clone_sheet(wb_iar.active, wb_pcv, "IAR")

    # ✅ CNRR SHEET (ONLY IF APPLICABLE)
    if voucher.has_cnrr:
        wb_cnrr = generate_cnrr_excel(voucher, administrator)
        clone_sheet(wb_cnrr.active, wb_pcv, "CNRR")

    # ================= OUTPUT =================
    buffer = BytesIO()
    wb_pcv.save(buffer)
    buffer.seek(0)

    filename = f"{voucher.pcv_no or voucher.uuid}.xlsx"

    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response


@login_required
@require_POST
@transaction.atomic
def reset_pettycash(request):

    if not request.user.has_role("Custodian"):
        return render(request, "403.html", status=403)

    if request.POST.get("confirm") != "RESET":
        messages.error(request, "Confirmation failed.")
        return redirect("users:dashboard_custodian")

    entity = request.user.entity

    # =========================================
    # 1. SOFT RESET VOUCHERS
    # =========================================
    PettyCashVoucher.objects.filter(entity=entity).update(
        status=VoucherStatus.DRAFT,
        is_posted_to_ledger=False,
        is_release_posted=False,
        is_liquidation_posted=False,
        is_replenished=False,
        replenishment=None,
        release_date=None,
        released_by=None,
        amount_liquidated=0,
        pr_no=None,
        iar_no=None,
    )

    # =========================================
    # 2. DELETE FINANCIAL DATA
    # =========================================
    LedgerEntry.objects.filter(fund__entity=entity).delete()
    Replenishment.objects.filter(fund__entity=entity).delete()

    # =========================================
    # 3. DELETE AUDIT LOGS (RECENT ACTIVITIES)
    # =========================================
    from audit.models import AuditLog
    audit_deleted, _ = AuditLog.objects.filter(entity=entity).delete()

    # =========================================
    # 4. DELETE NOTIFICATIONS
    # =========================================
    notifications_deleted, _ = Notification.objects.filter(
        user__entity=entity
    ).delete()

    # =========================================
    # 5. RESET FUND BALANCES
    # =========================================
    funds = PettyCashFund.objects.filter(entity=entity)

    for fund in funds:
        fund.current_balance = fund.fund_amount
        fund.save(update_fields=["current_balance"])

    # =========================================
    # SUCCESS MESSAGE
    # =========================================
    messages.success(
        request,
        f"Petty cash reset completed. "
        f"{audit_deleted} logs removed, "
        f"{notifications_deleted} notifications cleared."
    )

    return redirect("users:dashboard_custodian")

@login_required
def backup_pettycash(request):

    if not request.user.has_role("Custodian"):
        return render(request, "403.html", status=403)

    entity = request.user.entity

    vouchers = PettyCashVoucher.objects.filter(entity=entity)

    zip_buffer = BytesIO()

    with zipfile.ZipFile(zip_buffer, "w") as zip_file:

        for v in vouchers:

            content = f"""
                PCV: {v.pcv_no}
                Purpose: {v.purpose}
                Amount: {v.amount_requested}
                Date: {v.created_at}
                Status: {v.status}
            """

            filename = f"{v.pcv_no or v.uuid}.txt"
            zip_file.writestr(filename, content)

    zip_buffer.seek(0)

    response = HttpResponse(zip_buffer, content_type="application/zip")
    response["Content-Disposition"] = 'attachment; filename="pettycash_backup.zip"'

    return response

