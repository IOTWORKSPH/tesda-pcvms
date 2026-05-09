from datetime import date, datetime
from html import escape

from openpyxl.utils import get_column_letter


def worksheet_to_html(
    ws,
    css_class="excel-sheet",
    scale=1,
    max_row=None,
    max_col=None,
    html_overrides=None,
):
    bounds_max_row, bounds_max_col = _worksheet_bounds(ws)
    max_row = max_row or bounds_max_row
    max_col = max_col or bounds_max_col
    html_overrides = html_overrides or {}
    merge_starts, merge_covered = _merge_maps(ws, max_row=max_row, max_col=max_col)

    colgroup = []
    table_width = 0
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        width = _column_width_to_pixels(ws.column_dimensions[letter].width) * scale
        table_width += width
        colgroup.append(
            f'<col style="width:{width:.2f}px;">'
        )

    rows = []
    for row in range(1, max_row + 1):
        height = ws.row_dimensions[row].height or 15
        row_style = f' style="height:{height * scale:.2f}pt;"'
        cells = []

        for col in range(1, max_col + 1):
            if (row, col) in merge_covered:
                continue

            cell = ws.cell(row=row, column=col)
            merge_range = merge_starts.get((row, col))
            attrs = []

            if merge_range:
                rowspan = merge_range.max_row - merge_range.min_row + 1
                colspan = merge_range.max_col - merge_range.min_col + 1
                if rowspan > 1:
                    attrs.append(f'rowspan="{rowspan}"')
                if colspan > 1:
                    attrs.append(f'colspan="{colspan}"')

            style = _cell_style(ws, cell, merge_range, scale)
            if style:
                attrs.append(f'style="{style}"')

            content = html_overrides.get((row, col), _format_value(ws, cell))
            attr_text = " " + " ".join(attrs) if attrs else ""
            cells.append(f"<td{attr_text}>{content}</td>")

        rows.append(f"<tr{row_style}>{''.join(cells)}</tr>")

    return (
        f'<table class="{css_class}" style="width:{table_width:.2f}px;">'
        f"<colgroup>{''.join(colgroup)}</colgroup>"
        f"<tbody>{''.join(rows)}</tbody>"
        f"</table>"
    )


def worksheet_fit_scale(ws, orientation="portrait", margin_mm=8, max_row=None, max_col=None):
    bounds_max_row, bounds_max_col = _worksheet_bounds(ws)
    max_row = max_row or bounds_max_row
    max_col = max_col or bounds_max_col

    width_px = 0
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        width_px += _column_width_to_pixels(ws.column_dimensions[letter].width)

    height_pt = 0
    for row in range(1, max_row + 1):
        height_pt += ws.row_dimensions[row].height or 15

    sheet_width_mm = width_px * 0.264583
    sheet_height_mm = height_pt * 0.352778

    if orientation == "landscape":
        page_width_mm = 297 - (margin_mm * 2)
        page_height_mm = 210 - (margin_mm * 2)
    else:
        page_width_mm = 210 - (margin_mm * 2)
        page_height_mm = 297 - (margin_mm * 2)

    if sheet_width_mm <= 0 or sheet_height_mm <= 0:
        return 1

    return min(
        1,
        page_width_mm / sheet_width_mm,
        page_height_mm / sheet_height_mm,
    )


def _column_width_to_pixels(width):
    if width is None:
        width = 8.43

    # Excel stores column width as character units. This formula follows the
    # pixel conversion used by Excel closely enough for browser print layout.
    if width < 1:
        return int(width * 12)

    return int(width * 7 + 5)


def _worksheet_bounds(ws):
    max_row = 1
    max_col = 1

    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)

    for merged_range in ws.merged_cells.ranges:
        max_row = max(max_row, merged_range.max_row)
        max_col = max(max_col, merged_range.max_col)

    return max_row, max_col


def _merge_maps(ws, max_row=None, max_col=None):
    starts = {}
    covered = set()

    for merged_range in ws.merged_cells.ranges:
        if max_row and merged_range.min_row > max_row:
            continue

        if max_col and merged_range.min_col > max_col:
            continue

        starts[(merged_range.min_row, merged_range.min_col)] = merged_range

        range_max_row = min(merged_range.max_row, max_row or merged_range.max_row)
        range_max_col = min(merged_range.max_col, max_col or merged_range.max_col)

        for row in range(merged_range.min_row, range_max_row + 1):
            for col in range(merged_range.min_col, range_max_col + 1):
                if (row, col) != (merged_range.min_row, merged_range.min_col):
                    covered.add((row, col))

    return starts, covered


def _format_value(ws, cell):
    value = cell.value

    if value is None:
        return "&nbsp;"

    if isinstance(value, str) and value.startswith("="):
        value = _resolve_simple_formula(ws, value)
        if value is None:
            return "&nbsp;"

    if isinstance(value, (datetime, date)):
        return escape(value.strftime("%m-%d-%y"))

    if isinstance(value, (int, float)):
        if "0.00" in (cell.number_format or ""):
            return escape(f"{value:,.2f}")
        return escape(str(value))

    return escape(str(value)).replace("\n", "<br>")


def _resolve_simple_formula(ws, formula):
    reference = formula.lstrip("=").strip()

    if "!" in reference:
        return None

    try:
        referenced = ws[reference]
    except Exception:
        return None

    if referenced.value == formula:
        return None

    if isinstance(referenced.value, str) and referenced.value.startswith("="):
        return _resolve_simple_formula(ws, referenced.value)

    return referenced.value


def _cell_style(ws, cell, merge_range=None, scale=1):
    styles = [
        "box-sizing:border-box",
        f"padding:{2 * scale:.2f}px {3 * scale:.2f}px",
        "overflow:visible",
        "position:relative",
    ]

    font = cell.font
    if font:
        if font.name:
            styles.append(f"font-family:{font.name}, Arial, sans-serif")
        if font.sz:
            styles.append(f"font-size:{font.sz * scale:.2f}pt")
        if font.bold:
            styles.append("font-weight:700")
        if font.italic:
            styles.append("font-style:italic")
        if font.underline:
            styles.append("text-decoration:underline")
        color = _color_to_css(font.color)
        if color:
            styles.append(f"color:{color}")

    fill = cell.fill
    if fill and fill.fill_type:
        color = _color_to_css(fill.fgColor)
        if color:
            styles.append(f"background-color:{color}")

    alignment = cell.alignment
    if alignment:
        if alignment.horizontal:
            styles.append(f"text-align:{_horizontal_alignment(alignment.horizontal)}")
        if alignment.vertical:
            styles.append(f"vertical-align:{_vertical_alignment(alignment.vertical)}")
        if alignment.wrap_text:
            styles.append("white-space:pre-wrap")
        else:
            styles.append("white-space:pre")

    border_css = _border_style(ws, cell, merge_range, scale)
    styles.extend(border_css)

    return ";".join(styles)


def _border_style(ws, cell, merge_range=None, scale=1):
    if not merge_range:
        border = cell.border
        return [
            _side_to_css("left", border.left, scale),
            _side_to_css("right", border.right, scale),
            _side_to_css("top", border.top, scale),
            _side_to_css("bottom", border.bottom, scale),
        ]

    top = _first_side(
        ws.cell(merge_range.min_row, col).border.top
        for col in range(merge_range.min_col, merge_range.max_col + 1)
    )
    right = _first_side(
        ws.cell(row, merge_range.max_col).border.right
        for row in range(merge_range.min_row, merge_range.max_row + 1)
    )
    bottom = _first_side(
        ws.cell(merge_range.max_row, col).border.bottom
        for col in range(merge_range.min_col, merge_range.max_col + 1)
    )
    left = _first_side(
        ws.cell(row, merge_range.min_col).border.left
        for row in range(merge_range.min_row, merge_range.max_row + 1)
    )

    return [
        _side_to_css("left", left, scale),
        _side_to_css("right", right, scale),
        _side_to_css("top", top, scale),
        _side_to_css("bottom", bottom, scale),
    ]


def _first_side(sides):
    for side in sides:
        if side and side.style:
            return side
    return None


def _side_to_css(name, side, scale=1):
    if not side or not side.style:
        return f"border-{name}:0"

    widths = {
        "hair": "0.5px",
        "thin": "1px",
        "medium": "2px",
        "thick": "3px",
        "double": "3px",
    }
    line_styles = {
        "dashed": "dashed",
        "dashDot": "dashed",
        "dashDotDot": "dashed",
        "dotted": "dotted",
        "double": "double",
    }

    base_width = float(widths.get(side.style, "1px").replace("px", ""))
    width = f"{max(0.5, base_width * scale):.2f}px"
    line_style = line_styles.get(side.style, "solid")
    color = _color_to_css(side.color) or "#000000"
    return f"border-{name}:{width} {line_style} {color}"


def _color_to_css(color):
    if not color:
        return None

    if color.type == "rgb" and color.rgb:
        rgb = color.rgb[-6:]
        if rgb == "000000" and color.rgb in {"00000000", "000000"}:
            return None
        return f"#{rgb}"

    return None


def _horizontal_alignment(value):
    return {
        "center": "center",
        "right": "right",
        "left": "left",
        "fill": "left",
        "general": "left",
        "justify": "justify",
        "distributed": "center",
    }.get(value, value)


def _vertical_alignment(value):
    return {
        "center": "middle",
        "top": "top",
        "bottom": "bottom",
        "justify": "middle",
        "distributed": "middle",
    }.get(value, value)
