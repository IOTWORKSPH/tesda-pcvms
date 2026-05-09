# pettycash/services/excel/pr_excel.py

from copy import copy

from openpyxl import load_workbook
from django.conf import settings
import os

from pettycash.services.excel.document_numbers import preliminary_document_number



def generate_pr_excel(voucher, administrator):
    """
    Populate PR template using exact TESDA format
    """

    template_path = os.path.join(
        settings.BASE_DIR,
        "pettycash",
        "templates",
        "excel_templates",
        "PR_template.xlsx"
    )

    wb = load_workbook(template_path)
    ws = wb.active  # use first sheet

    # =====================================================
    # HEADER DETAILS
    # =====================================================

    document_date = (
        voucher.purchase_date
        or (voucher.release_date.date() if voucher.release_date else None)
        or voucher.created_at.date()
    )

    # PR Number
    ws["C8"] = f"PR No.: {preliminary_document_number('PR', document_date)}"

    # Date
    ws["E8"] = document_date.strftime("%B %d, %Y")

    # Fund Source
    ws["F7"] = voucher.fund.fund_cluster.code if voucher.fund else ""

    # Office / Section
    ws["A8"] = f"Office/Section: {voucher.entity.name}"

    # =====================================================
    # ITEMS TABLE
    # START ROW = 12
    # =====================================================

    start_row = 12
    current_row = start_row

    total_amount = 0

    items = voucher.items.all()

    for index, item in enumerate(items, start=1):

        # A = Series #
        ws[f"A{current_row}"] = index

        # B = Unit
        ws[f"B{current_row}"] = item.unit

        # C = Description
        ws[f"C{current_row}"] = item.description

        # D = Quantity
        ws[f"D{current_row}"] = float(item.quantity)

        # E = Unit Cost
        ws[f"E{current_row}"] = float(item.unit_cost)

        # F = Total Cost
        line_total = float(item.quantity * item.unit_cost)
        ws[f"F{current_row}"] = line_total

        total_amount += line_total
        current_row += 1

    # =====================================================
    # TOTAL
    # =====================================================
    ws["F24"] = total_amount
    # =====================================================
    # PURPOSED
    # =====================================================
    ws["B25"] = voucher.purpose

    # =====================================================
    # SIGNATORIES
    # =====================================================

    # Requestor
    ws["C28"] = voucher.requester.get_full_name().upper()
    ws["C29"] = getattr(voucher.requester, "position", "")
    for cell_ref in ("C28", "C29"):
        alignment = copy(ws[cell_ref].alignment)
        alignment.horizontal = "left"
        ws[cell_ref].alignment = alignment

    # Administrator (Approving Officer)
    if administrator:
        ws["D28"] = administrator.get_full_name().upper()
        ws["D29"] = getattr(administrator, "position", "")

    return wb
