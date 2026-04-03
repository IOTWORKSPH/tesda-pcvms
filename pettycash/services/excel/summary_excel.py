# pettycash/services/excel/summary_excel.py

from decimal import Decimal
from openpyxl.styles import Font


def _get_voucher_amount(voucher):
    actual_amount = getattr(voucher, "actual_amount", None)
    if actual_amount is not None:
        return actual_amount

    if getattr(voucher, "amount_liquidated", None):
        return voucher.amount_liquidated

    return voucher.amount_requested or Decimal("0.00")


def generate_summary(wb, context, styles):
    fund = context["fund"]
    vouchers = context["vouchers"]
    total = context.get("replenishment_amount", context.get("total", Decimal("0.00")))
    generated_date = context["generated_date"]

    ws = wb.create_sheet("Summary")

    bold = styles["bold"]
    center = styles["center"]
    right = styles["right"]
    border = styles["border"]
    wrap = styles["wrap"]

    row = 1

    # =====================================================
    # HEADER
    # =====================================================
    ws.merge_cells("A1:D1")
    ws["A1"] = "Republic of the Philippines"
    ws["A1"].alignment = center
    row += 1

    ws.merge_cells("A2:D2")
    ws["A2"] = "TECHNICAL EDUCATION AND SKILLS DEVELOPMENT AUTHORITY"
    ws["A2"].font = bold
    ws["A2"].alignment = center
    row += 1

    ws.merge_cells("A3:D3")
    ws["A3"] = fund.entity.name.upper()
    ws["A3"].font = bold
    ws["A3"].alignment = center
    row += 1

    ws.merge_cells("A4:D4")
    ws["A4"] = fund.entity.address or ""
    ws["A4"].alignment = center
    row += 2

    # =====================================================
    # TITLE
    # =====================================================
    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1).value = "PETTY CASH REPLENISHMENT REPORT"
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    ws.cell(row=row, column=1).alignment = center
    row += 2

    # =====================================================
    # TABLE HEADER
    # =====================================================
    headers = ["Date", "Petty Cash Voucher No.", "Particulars", "Amount"]

    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=idx)
        cell.value = header
        cell.font = bold
        cell.alignment = center
        cell.border = border

    row += 1

    # =====================================================
    # TABLE DATA
    # =====================================================
    for voucher in vouchers:
        amount = _get_voucher_amount(voucher)

        ws.cell(row=row, column=1).value = (
            voucher.purchase_date.strftime("%B %d, %Y")
            if voucher.purchase_date else ""
        )
        ws.cell(row=row, column=1).alignment = center

        ws.cell(row=row, column=2).value = getattr(voucher, "report_pcv_no", voucher.pcv_no)
        ws.cell(row=row, column=2).alignment = center

        ws.cell(row=row, column=3).value = (
            voucher.expense_category.name if voucher.expense_category else ""
        )
        ws.cell(row=row, column=3).alignment = wrap

        ws.cell(row=row, column=4).value = float(amount)
        ws.cell(row=row, column=4).alignment = right
        ws.cell(row=row, column=4).number_format = '#,##0.00'

        for col in range(1, 5):
            ws.cell(row=row, column=col).border = border

        row += 1

    # =====================================================
    # TOTAL ROW
    # =====================================================
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1).value = "TOTAL"
    ws.cell(row=row, column=1).alignment = right
    ws.cell(row=row, column=1).font = bold
    ws.cell(row=row, column=1).border = border

    ws.cell(row=row, column=4).value = float(total)
    ws.cell(row=row, column=4).alignment = right
    ws.cell(row=row, column=4).font = bold
    ws.cell(row=row, column=4).border = border
    ws.cell(row=row, column=4).number_format = '#,##0.00'

    row += 3

    # =====================================================
    # CERTIFICATION
    # =====================================================
    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1).value = "CERTIFICATION"
    ws.cell(row=row, column=1).font = bold
    ws.cell(row=row, column=1).alignment = center
    row += 2

    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1).value = "I hereby certify to the correctness of the above information."
    ws.cell(row=row, column=1).alignment = center
    row += 4

    # =====================================================
    # SIGNATURE BLOCK
    # =====================================================
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1).value = fund.custodian.get_full_name().upper()
    ws.cell(row=row, column=1).font = bold
    ws.cell(row=row, column=1).alignment = center

    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    ws.cell(row=row, column=3).value = generated_date.strftime("%B %d, %Y")
    ws.cell(row=row, column=3).font = bold
    ws.cell(row=row, column=3).alignment = center
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1).value = "Petty Cash Custodian"
    ws.cell(row=row, column=1).alignment = center

    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    ws.cell(row=row, column=3).value = "Date"
    ws.cell(row=row, column=3).alignment = center

    # =====================================================
    # COLUMN WIDTHS
    # =====================================================
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 18

    # =====================================================
    # PAGE SETUP
    # =====================================================
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.print_area = f"A1:D{row}"