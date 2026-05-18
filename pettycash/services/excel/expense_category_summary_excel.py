from decimal import Decimal

from openpyxl.styles import Font


def generate_expense_category_summary(wb, context, styles):
    fund = context["fund"]
    summary_rows = context.get("expense_category_summary", [])
    total = context.get("replenishment_amount", context.get("total", Decimal("0.00")))

    ws = wb.create_sheet("Expense Category Summary")

    bold = styles["bold"]
    center = styles["center"]
    right = styles["right"]
    border = styles["border"]
    wrap = styles["wrap"]

    row = 1

    ws.merge_cells("A1:C1")
    ws["A1"] = "Republic of the Philippines"
    ws["A1"].alignment = center
    row += 1

    ws.merge_cells("A2:C2")
    ws["A2"] = "TECHNICAL EDUCATION AND SKILLS DEVELOPMENT AUTHORITY"
    ws["A2"].font = bold
    ws["A2"].alignment = center
    row += 1

    ws.merge_cells("A3:C3")
    ws["A3"] = fund.entity.name.upper()
    ws["A3"].font = bold
    ws["A3"].alignment = center
    row += 1

    ws.merge_cells("A4:C4")
    ws["A4"] = fund.entity.address or ""
    ws["A4"].alignment = center
    row += 2

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1).value = "SUMMARY OF EXPENSES PER EXPENSE CATEGORY"
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    ws.cell(row=row, column=1).alignment = center
    row += 2

    headers = ["UACS Code", "Expense Category", "Amount"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=idx)
        cell.value = header
        cell.font = bold
        cell.alignment = center
        cell.border = border

    row += 1

    for summary in summary_rows:
        ws.cell(row=row, column=1).value = summary["code"]
        ws.cell(row=row, column=1).alignment = center

        ws.cell(row=row, column=2).value = summary["name"]
        ws.cell(row=row, column=2).alignment = wrap

        ws.cell(row=row, column=3).value = float(summary["amount"])
        ws.cell(row=row, column=3).alignment = right
        ws.cell(row=row, column=3).number_format = '#,##0.00'

        for col in range(1, 4):
            ws.cell(row=row, column=col).border = border

        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1).value = "TOTAL"
    ws.cell(row=row, column=1).alignment = right
    ws.cell(row=row, column=1).font = bold
    ws.cell(row=row, column=1).border = border

    ws.cell(row=row, column=3).value = float(total)
    ws.cell(row=row, column=3).alignment = right
    ws.cell(row=row, column=3).font = bold
    ws.cell(row=row, column=3).border = border
    ws.cell(row=row, column=3).number_format = '#,##0.00'

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 18

    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.print_area = f"A1:C{row}"
