# pettycash/services/excel/appendix_49_excel.py

from decimal import Decimal
from openpyxl.styles import Alignment


def _get_period_text(context):
    date_from = context.get("date_from") or context.get("period_start")
    date_to = context.get("date_to") or context.get("period_end")

    if date_from and date_to:
        return f"{date_from.strftime('%B %d, %Y')} - {date_to.strftime('%B %d, %Y')}"
    if date_from:
        return date_from.strftime("%B %d, %Y")
    if date_to:
        return date_to.strftime("%B %d, %Y")
    return "No Transactions Available"


def _get_voucher_amount(voucher):
    actual_amount = getattr(voucher, "actual_amount", None)
    if actual_amount is not None:
        return actual_amount

    if getattr(voucher, "amount_liquidated", None):
        return voucher.amount_liquidated

    return voucher.amount_requested or Decimal("0.00")


def generate_appendix_49(wb, context, styles):
    fund = context["fund"]
    vouchers = context["vouchers"]
    total = context.get("replenishment_amount", context.get("total", Decimal("0.00")))
    generated_date = context["generated_date"]

    report_number = context.get("report_number", "") or ""
    sheet_number = context.get("sheet_number", 1) or 1
    period_text = _get_period_text(context)

    ws = wb.active
    ws.title = "Appendix 49"

    bold = styles["bold"]
    bold_large = styles["bold_large"]
    center = styles["center"]
    right = styles["right"]
    border = styles["border"]
    wrap = styles["wrap"]

    row = 1

    # =====================================================
    # APPENDIX LABEL
    # =====================================================
    ws.merge_cells("A1:D1")
    ws["A1"] = "Appendix 49"
    ws["A1"].alignment = Alignment(horizontal="right")
    row += 1

    # =====================================================
    # TITLE
    # =====================================================
    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1).value = "REPORT ON PAID PETTY CASH VOUCHERS"
    ws.cell(row=row, column=1).font = bold_large
    ws.cell(row=row, column=1).alignment = center
    row += 1

    # =====================================================
    # PERIOD COVERED
    # =====================================================
    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1).value = f"Period Covered: {period_text}"
    ws.cell(row=row, column=1).alignment = center
    row += 2

    # =====================================================
    # ENTITY + REPORT INFO
    # =====================================================
    ws.cell(row=row, column=1).value = "Entity Name:"
    ws.cell(row=row, column=1).font = bold
    ws.cell(row=row, column=2).value = fund.entity.name

    ws.cell(row=row, column=3).value = "Report No:"
    ws.cell(row=row, column=3).font = bold
    ws.cell(row=row, column=4).value = report_number
    row += 1

    ws.cell(row=row, column=1).value = "Fund Cluster:"
    ws.cell(row=row, column=1).font = bold
    ws.cell(row=row, column=2).value = fund.fund_cluster.code

    ws.cell(row=row, column=3).value = "Sheet No:"
    ws.cell(row=row, column=3).font = bold
    ws.cell(row=row, column=4).value = sheet_number
    row += 2

    # =====================================================
    # TABLE HEADER
    # =====================================================
    headers = ["Date", "Petty Cash Voucher No.", "Particulars", "Amount"]

    for idx, header in enumerate(headers, start=1):
        ws.cell(row=row, column=idx).value = header
        ws.cell(row=row, column=idx).font = bold
        ws.cell(row=row, column=idx).alignment = center
        ws.cell(row=row, column=idx).border = border

    row += 1

    # =====================================================
    # TABLE DATA
    # =====================================================
    for voucher in vouchers:
        amount = _get_voucher_amount(voucher)

        ws.cell(row=row, column=1).value = (
            voucher.purchase_date.strftime("%m-%d-%Y")
            if voucher.purchase_date else ""
        )

        ws.cell(row=row, column=2).value = getattr(voucher, "report_pcv_no", voucher.pcv_no)
        ws.cell(row=row, column=2).alignment = center

        ws.cell(row=row, column=3).value = (
            voucher.expense_category.name if voucher.expense_category else ""
        )
        ws.cell(row=row, column=3).alignment = wrap

        ws.cell(row=row, column=4).value = float(amount)
        ws.cell(row=row, column=4).alignment = right
        ws.cell(row=row, column=4).number_format = '#,##0.00'

        for col in range(1, 5):
            ws.cell(row=row, column=col).border = border

        row += 1

    # =====================================================
    # TOTAL ROW
    # =====================================================
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1).value = "TOTAL"
    ws.cell(row=row, column=1).alignment = right
    ws.cell(row=row, column=1).font = bold
    ws.cell(row=row, column=1).border = border

    ws.cell(row=row, column=4).value = float(total)
    ws.cell(row=row, column=4).alignment = right
    ws.cell(row=row, column=4).font = bold
    ws.cell(row=row, column=4).border = border
    ws.cell(row=row, column=4).number_format = '#,##0.00'

    row += 3

    # =====================================================
    # CERTIFICATION
    # =====================================================
    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1).value = "C E R T I F I C A T I O N"
    ws.cell(row=row, column=1).font = bold
    ws.cell(row=row, column=1).alignment = center
    row += 2

    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1).value = "I hereby certify to the correctness of the above information."
    ws.cell(row=row, column=1).alignment = center
    row += 4

    # =====================================================
    # SIGNATURE BLOCK
    # =====================================================
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1).value = fund.custodian.get_full_name().upper()
    ws.cell(row=row, column=1).font = bold
    ws.cell(row=row, column=1).alignment = center

    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    ws.cell(row=row, column=3).value = generated_date.strftime("%B %d, %Y")
    ws.cell(row=row, column=3).font = bold
    ws.cell(row=row, column=3).alignment = center
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1).value = "Petty Cash Custodian"
    ws.cell(row=row, column=1).alignment = center

    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    ws.cell(row=row, column=3).value = f"Date: {generated_date.strftime('%B %d, %Y')}"
    ws.cell(row=row, column=3).alignment = center

    # =====================================================
    # COLUMN WIDTHS
    # =====================================================
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 18

    # =====================================================
    # PAGE SETUP
    # =====================================================
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.print_area = f"A1:D{row}"