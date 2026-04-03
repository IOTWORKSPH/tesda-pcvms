# pettycash/services/excel/appendix_51_excel.py

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def generate_appendix_51(wb, context, styles):
    fund = context["fund"]
    register_rows = context["register_rows"]
    expense_categories = context["expense_categories"]
    category_totals = context["category_totals"]
    total = context["total"]
    generated_date = context["generated_date"]

    ws = wb.create_sheet("Appendix 51")

    bold = styles["bold"]
    bold_large = styles["bold_large"]
    center = styles["center"]
    right = styles["right"]
    wrap = styles["wrap"]
    border = styles["border"]

    row = 1
    total_columns = 6 + len(expense_categories)

    # =====================================================
    # APPENDIX LABEL
    # =====================================================
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_columns)
    ws.cell(row=row, column=1).value = "Appendix 51"
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
    row += 1

    # =====================================================
    # TITLE
    # =====================================================
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_columns)
    ws.cell(row=row, column=1).value = "PETTY CASH FUND REGISTER"
    ws.cell(row=row, column=1).font = bold_large
    ws.cell(row=row, column=1).alignment = center
    row += 2

    # =====================================================
    # HEADER INFO
    # =====================================================
    ws.cell(row=row, column=1).value = "Department/Agency:"
    ws.cell(row=row, column=1).font = bold
    ws.cell(row=row, column=2).value = "TESDA"

    ws.cell(row=row, column=5).value = "Petty Cash Fund Custodian:"
    ws.cell(row=row, column=5).font = bold
    ws.cell(row=row, column=6).value = fund.custodian.get_full_name()
    row += 1

    ws.cell(row=row, column=1).value = "Sub-Office/District/Division:"
    ws.cell(row=row, column=1).font = bold
    ws.cell(row=row, column=2).value = fund.entity.name

    ws.cell(row=row, column=5).value = "Fund Cluster:"
    ws.cell(row=row, column=5).font = bold
    ws.cell(row=row, column=6).value = fund.fund_cluster.code
    row += 1

    ws.cell(row=row, column=1).value = "Municipality/City/Province:"
    ws.cell(row=row, column=1).font = bold
    ws.cell(row=row, column=2).value = fund.entity.address or "-"

    ws.cell(row=row, column=5).value = "Sheet No:"
    ws.cell(row=row, column=5).font = bold
    ws.cell(row=row, column=6).value = 1
    row += 2

    # =====================================================
    # COMPLEX HEADER
    # =====================================================
    header_start = row

    ws.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=1)
    ws.cell(row=row, column=1).value = "Date"

    ws.merge_cells(start_row=row, start_column=2, end_row=row + 2, end_column=2)
    ws.cell(row=row, column=2).value = "PCV / Check No."

    ws.merge_cells(start_row=row, start_column=3, end_row=row + 2, end_column=3)
    ws.cell(row=row, column=3).value = "Particulars"

    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    ws.cell(row=row, column=4).value = "PETTY CASH FUND"

    breakdown_start = 7
    breakdown_end = total_columns

    ws.merge_cells(start_row=row, start_column=breakdown_start, end_row=row, end_column=breakdown_end)
    ws.cell(row=row, column=breakdown_start).value = "BREAKDOWN OF PAYMENTS"

    row += 1

    ws.cell(row=row, column=4).value = "Receipts"
    ws.cell(row=row, column=5).value = "Payments"
    ws.cell(row=row, column=6).value = "Balance"

    col = 7
    for category in expense_categories:
        ws.cell(row=row, column=col).value = category.name
        ws.cell(row=row, column=col).alignment = wrap
        col += 1

    row += 1

    ws.cell(row=row, column=4).value = "( + )"
    ws.cell(row=row, column=5).value = "( - )"

    for r in range(header_start, header_start + 3):
        ws.row_dimensions[r].height = 35
        for c in range(1, total_columns + 1):
            ws.cell(row=r, column=c).font = bold
            ws.cell(row=r, column=c).alignment = center
            ws.cell(row=r, column=c).border = border

    row += 1

    # =====================================================
    # DATA ROWS
    # =====================================================
    for entry in register_rows:
        ws.cell(row=row, column=1).value = (
            entry["date"].strftime("%B %d, %Y")
            if entry.get("date") else ""
        )
        ws.cell(row=row, column=2).value = entry.get("reference") or ""

        ws.cell(row=row, column=3).value = entry.get("particulars") or ""
        ws.cell(row=row, column=3).alignment = wrap

        if entry.get("receipt") is not None:
            ws.cell(row=row, column=4).value = float(entry["receipt"])
            ws.cell(row=row, column=4).number_format = '#,##0.00'
        else:
            ws.cell(row=row, column=4).value = ""

        if entry.get("payment") is not None:
            ws.cell(row=row, column=5).value = float(entry["payment"])
            ws.cell(row=row, column=5).number_format = '#,##0.00'
        else:
            ws.cell(row=row, column=5).value = ""

        if entry.get("balance") is not None:
            ws.cell(row=row, column=6).value = float(entry["balance"])
            ws.cell(row=row, column=6).number_format = '#,##0.00'
        else:
            ws.cell(row=row, column=6).value = ""

        ws.cell(row=row, column=4).alignment = right
        ws.cell(row=row, column=5).alignment = right
        ws.cell(row=row, column=6).alignment = right

        col = 7
        for category in expense_categories:
            amount = entry["breakdown"].get(category.id)
            if amount is not None:
                ws.cell(row=row, column=col).value = float(amount)
                ws.cell(row=row, column=col).number_format = '#,##0.00'
            else:
                ws.cell(row=row, column=col).value = ""
            ws.cell(row=row, column=col).alignment = right
            col += 1

        for c in range(1, total_columns + 1):
            ws.cell(row=row, column=c).border = border

        row += 1

    # =====================================================
    # TOTAL ROW
    # =====================================================
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1).value = "TOTAL"
    ws.cell(row=row, column=1).font = bold
    ws.cell(row=row, column=1).alignment = right
    ws.cell(row=row, column=1).border = border

    ws.cell(row=row, column=5).value = float(total)
    ws.cell(row=row, column=5).font = bold
    ws.cell(row=row, column=5).alignment = right
    ws.cell(row=row, column=5).border = border
    ws.cell(row=row, column=5).number_format = '#,##0.00'

    ws.cell(row=row, column=6).border = border

    col = 7
    for category in expense_categories:
        ws.cell(row=row, column=col).value = float(category_totals.get(category.id, 0))
        ws.cell(row=row, column=col).font = bold
        ws.cell(row=row, column=col).alignment = right
        ws.cell(row=row, column=col).border = border
        ws.cell(row=row, column=col).number_format = '#,##0.00'
        col += 1

    row += 4

    # =====================================================
    # SIGNATURE
    # =====================================================
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=total_columns - 2)
    ws.cell(row=row, column=3).value = fund.custodian.get_full_name().upper()
    ws.cell(row=row, column=3).font = bold
    ws.cell(row=row, column=3).alignment = center
    row += 1

    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=total_columns - 2)
    ws.cell(row=row, column=3).value = "(Signature over Printed)"
    ws.cell(row=row, column=3).alignment = center
    row += 1

    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=total_columns - 2)
    ws.cell(row=row, column=3).value = "Name Petty Cash Fund Custodian"
    ws.cell(row=row, column=3).alignment = center
    row += 1

    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=total_columns - 2)
    ws.cell(row=row, column=3).value = generated_date.strftime("%B %d, %Y")
    ws.cell(row=row, column=3).alignment = center

    # =====================================================
    # COLUMN WIDTHS
    # =====================================================
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12

    col = 7
    for _ in expense_categories:
        ws.column_dimensions[get_column_letter(col)].width = 16
        col += 1

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1