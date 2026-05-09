from openpyxl import load_workbook
from django.conf import settings
import os
from users.models import User

from pettycash.services.excel.document_numbers import preliminary_document_number


def generate_iar_excel(voucher):

    # =====================================================
    # LOAD TEMPLATE
    # =====================================================
    template_path = os.path.join(
        settings.BASE_DIR,
        "pettycash",
        "templates",
        "excel_templates",
        "IAR_template.xlsx"
    )

    wb = load_workbook(template_path)
    ws = wb.active

    entity = voucher.entity
    document_date = (
        voucher.purchase_date
        or (voucher.release_date.date() if voucher.release_date else None)
        or voucher.created_at.date()
    )

    # =====================================================
    # HEADER
    # =====================================================

    # Entity Name
    ws["B5"] = entity.name

    # Fund Cluster
    ws["E5"] = voucher.fund.fund_cluster.code if voucher.fund else ""

    # Supplier
    ws["B7"] = voucher.supplier.name if voucher.supplier else ""

    # IAR Number
    ws["E7"] = preliminary_document_number("IAR", document_date)

    # Invoice Number
    ws["E9"] = voucher.official_receipt_number or ""

    # Date Purchased
    ws["E10"] = document_date.strftime("%B %d, %Y")

    # =====================================================
    # ITEMS TABLE
    # =====================================================

    start_row = 13

    for i, item in enumerate(voucher.items.all(), start=0):
        row = start_row + i

        ws[f"A{row}"] = i + 1  # Series Number
        ws[f"B{row}"] = item.description
        ws[f"D{row}"] = item.unit
        ws[f"E{row}"] = float(item.quantity)

    # =====================================================
    # DATES (BOTTOM)
    # =====================================================

    formatted_date = document_date.strftime("%B %d, %Y")

    ws["B23"] = formatted_date
    ws["D23"] = formatted_date

    # =====================================================
    # INSPECTION TEAM (GROUP: Inspection)
    # =====================================================

    inspectors = User.objects.filter(
        entity=entity,
        groups__name="Inspection",
        is_active=True
    )

    inspector_list = list(inspectors)

    # First Inspector
    if len(inspector_list) >= 1:
        ws["A30"] = inspector_list[0].get_full_name().upper()
        ws["A31"] = getattr(inspector_list[0], "position", "")

    # Second Inspector
    if len(inspector_list) >= 2:
        ws["A34"] = inspector_list[1].get_full_name().upper()
        ws["A35"] = getattr(inspector_list[1], "position", "")

    # =====================================================
    # SUPPLY OFFICER (GROUP: Supply)
    # =====================================================

    supply = User.objects.filter(
        entity=entity,
        groups__name="Supply",
        is_active=True
    ).first()

    if supply:
        ws["C30"] = supply.get_full_name().upper()

    return wb
