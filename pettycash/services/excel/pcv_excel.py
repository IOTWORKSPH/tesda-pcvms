# pettycash/services/excel/pcv_excel.py

from openpyxl import load_workbook
from django.conf import settings
import os

from pettycash.services.excel.document_numbers import preliminary_document_number


def generate_pcv_excel(voucher, entity, administrator, custodian):

    # ✅ Load template file
    template_path = os.path.join(
        settings.BASE_DIR,
        "pettycash",
        "templates",
        "excel_templates",
        "PCV_template.xlsx"
    )

    wb = load_workbook(template_path)
    ws = wb.active

    # =====================================================
    # HEADER
    # =====================================================

    document_date = (
        voucher.purchase_date
        or (voucher.release_date.date() if voucher.release_date else None)
        or voucher.created_at.date()
    )

    ws["N2"] = preliminary_document_number("PCV", document_date)
    ws["N4"] = document_date.strftime("%B %d, %Y")

    ws["C4"] = entity.name
    ws["N8"] = voucher.fund.responsibility_center.code

    # =====================================================
    # PAYEE
    # =====================================================

    ws["C8"] = voucher.requester.get_full_name().upper()
    ws["C9"] = getattr(voucher.requester, "office", "")

    # =====================================================
    # REQUEST DETAILS
    # =====================================================

    ws["C14"] = voucher.expense_category.name if voucher.expense_category else ""
    ws["D14"] = float(voucher.amount_requested)

    # =====================================================
    # LIQUIDATION
    # =====================================================

    ws["N13"] = float(voucher.amount_requested)

    paid_amount = voucher.amount_liquidated or voucher.amount_requested
    ws["N14"] = float(paid_amount or 0)

    ws["N15"] = voucher.official_receipt_number or ""

    if voucher.variance_type == "EXCESS":
        ws["N17"] = float(voucher.variance_display or 0)

    elif voucher.variance_type == "SHORTAGE":
        ws["N18"] = float(voucher.variance_display or 0)

    # =====================================================
    # SIGNATORIES
    # =====================================================

    ws["C22"] = voucher.requester.get_full_name().upper()

    if administrator:
        ws["C27"] = administrator.get_full_name().upper()

    if custodian:
        ws["J27"] = custodian.get_full_name().upper()

    ws["C31"] = document_date.strftime("%B %d, %Y")
    ws["J31"] = document_date.strftime("%B %d, %Y")

    ws["C35"] = custodian.get_full_name().upper() if custodian else ""

    ws["C41"] = voucher.requester.get_full_name().upper()
    ws["J41"] = voucher.requester.get_full_name().upper()

    ws["C45"] = document_date.strftime("%B %d, %Y")
    ws["J45"] = document_date.strftime("%B %d, %Y")

    return wb
