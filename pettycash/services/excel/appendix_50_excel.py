# pettycash/services/excel/appendix_50_excel.py

from openpyxl.styles import Alignment


def _get_period_text(context):
    date_from = context.get("date_from") or context.get("period_start")
    date_to = context.get("date_to") or context.get("period_end")

    if date_from and date_to:
        return f"{date_from.strftime('%B %d, %Y')} - {date_to.strftime('%B %d, %Y')}"
    if date_from:
        return date_from.strftime("%B %d, %Y")
    if date_to:
        return date_to.strftime("%B %d, %Y")
    return "No Transactions Available"


def generate_appendix_50(wb, context, styles):
    fund = context["fund"]
    records = context["records"]
    generated_date = context["generated_date"]
    period_text = _get_period_text(context)

    ws = wb.create_sheet("Appendix 50")

    bold = styles["bold"]
    bold_large = styles["bold_large"]
    center = styles["center"]
    right = styles["right"]
    border = styles["border"]
    wrap = styles["wrap"]

    row = 1

    # =====================================================
    # APPENDIX LABEL
    # =====================================================
    ws.merge_cells("A1:G1")
    ws["A1"] = "Appendix 50"
    ws["A1"].alignment = Alignment(horizontal="right")
    row += 1

    # =====================================================
    # TITLE
    # =====================================================
    ws.merge_cells(f"A{row}:G{row}")
    ws.cell(row=row, column=1).value = "PETTY CASH FUND RECORD"
    ws.cell(row=row, column=1).font = bold_large
    ws.cell(row=row, column=1).alignment = center
    row += 2

    # =====================================================
    # ENTITY + FUND
    # =====================================================
    ws.cell(row=row, column=1).value = "Entity Name :"
    ws.cell(row=row, column=1).font = bold
    ws.cell(row=row, column=2).value = fund.entity.name
    row += 1

    ws.cell(row=row, column=1).value = "Fund Cluster :"
    ws.cell(row=row, column=1).font = bold
    ws.cell(row=row, column=2).value = fund.fund_cluster.code
    row += 2

    # =====================================================
    # CUSTODIAN HEADER
    # =====================================================
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)

    ws.cell(row=row, column=1).value = fund.custodian.get_full_name().upper()
    ws.cell(row=row, column=1).alignment = center
    ws.cell(row=row, column=1).font = bold

    ws.cell(row=row, column=3).value = fund.custodian.position or "-"
    ws.cell(row=row, column=3).alignment = center
    ws.cell(row=row, column=3).font = bold

    ws.cell(row=row, column=5).value = fund.custodian.office or fund.entity.name
    ws.cell(row=row, column=5).alignment = center
    ws.cell(row=row, column=5).font = bold
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1).value = "Petty Cash Fund Custodian"
    ws.cell(row=row, column=1).alignment = center

    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    ws.cell(row=row, column=3).value = "Official Designation"
    ws.cell(row=row, column=3).alignment = center

    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
    ws.cell(row=row, column=5).value = "Station"
    ws.cell(row=row, column=5).alignment = center
    row += 2

    # =====================================================
    # TABLE HEADER
    # =====================================================
    headers = [
        "Date",
        "Reference",
        "Payee",
        "Particulars",
        "Receipts",
        "Disbursements",
        "Balance",
    ]

    for idx, header in enumerate(headers, start=1):
        ws.cell(row=row, column=idx).value = header
        ws.cell(row=row, column=idx).font = bold
        ws.cell(row=row, column=idx).alignment = center
        ws.cell(row=row, column=idx).border = border

    row += 1

    # =====================================================
    # DATA ROWS
    # =====================================================
    for entry in records:
        entry_date = entry.get("date")
        received = entry.get("received")
        disbursement = entry.get("disbursement")
        balance = entry.get("balance")

        ws.cell(row=row, column=1).value = (
            entry_date.strftime("%B %d, %Y") if entry_date else ""
        )

        ws.cell(row=row, column=2).value = entry.get("reference") or ""
        ws.cell(row=row, column=3).value = entry.get("payee") or ""
        ws.cell(row=row, column=4).value = entry.get("particulars") or ""
        ws.cell(row=row, column=4).alignment = wrap

        if received is not None:
            ws.cell(row=row, column=5).value = float(received)
            ws.cell(row=row, column=5).number_format = '#,##0.00'
        else:
            ws.cell(row=row, column=5).value = ""

        if disbursement is not None:
            ws.cell(row=row, column=6).value = float(disbursement)
            ws.cell(row=row, column=6).number_format = '#,##0.00'
        else:
            ws.cell(row=row, column=6).value = ""

        if balance is not None:
            ws.cell(row=row, column=7).value = float(balance)
            ws.cell(row=row, column=7).number_format = '#,##0.00'
        else:
            ws.cell(row=row, column=7).value = ""

        ws.cell(row=row, column=5).alignment = right
        ws.cell(row=row, column=6).alignment = right
        ws.cell(row=row, column=7).alignment = right

        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border

        row += 1

    # =====================================================
    # PERIOD DISPLAY
    # =====================================================
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=1).value = f"Period Covered: {period_text}"
    ws.cell(row=row, column=1).alignment = center
    row += 3

    # =====================================================
    # CERTIFICATION
    # =====================================================
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=1).value = "C E R T I F I C A T I O N"
    ws.cell(row=row, column=1).font = bold
    ws.cell(row=row, column=1).alignment = center
    row += 2

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=1).value = (
        f"I hereby certify that the foregoing is a correct and complete record of "
        f"all cash advances received and disbursements made by me in my capacity "
        f"as Petty Cash Fund Custodian of {fund.entity.name} during the period from "
        f"{period_text}, inclusive, as indicated in the corresponding columns."
    )
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", wrap_text=True)
    row += 4

    # =====================================================
    # SIGNATURE BLOCK
    # =====================================================
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1).value = fund.custodian.get_full_name().upper()
    ws.cell(row=row, column=1).font = bold
    ws.cell(row=row, column=1).alignment = center

    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
    ws.cell(row=row, column=5).value = generated_date.strftime("%B %d, %Y")
    ws.cell(row=row, column=5).font = bold
    ws.cell(row=row, column=5).alignment = center
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1).value = "Name and Signature of Petty Cash Fund Custodian"
    ws.cell(row=row, column=1).alignment = center

    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
    ws.cell(row=row, column=5).value = "Date"
    ws.cell(row=row, column=5).alignment = center

    # =====================================================
    # COLUMN WIDTHS
    # =====================================================
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16

    # =====================================================
    # PAGE SETUP
    # =====================================================
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.print_area = f"A1:G{row}"