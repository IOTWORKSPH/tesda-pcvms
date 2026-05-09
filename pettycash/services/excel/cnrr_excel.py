# pettycash/services/excel/cnrr_excel.py

from openpyxl import load_workbook
from django.conf import settings
import os


def generate_cnrr_excel(voucher, administrator):

    template_path = os.path.join(
        settings.BASE_DIR,
        "pettycash",
        "templates",
        "excel_templates",
        "CNRR_template.xlsx"
    )

    wb = load_workbook(template_path)
    ws = wb.active

    requester = voucher.requester
    entity = voucher.entity

    # ================= HEADER =================
    fullname = requester.get_full_name().upper()
    employee_number = getattr(requester, "employee_number", "") or ""
    office = getattr(requester, "office", "") or "TESDA"
    division = entity.name if entity else ""

    ws["C5"] = fullname
    ws["N5"] = fullname

    ws["I5"] = employee_number
    ws["T5"] = employee_number

    ws["C6"] = office
    ws["N6"] = office

    ws["C7"] = division
    ws["N7"] = division

    # ================= PARTICULARS =================
    for row in range(9, 18):
        ws[f"A{row}"] = ""
        ws[f"I{row}"] = None
        ws[f"L{row}"] = ""
        ws[f"T{row}"] = None

    total_amount = 0

    for row, item in zip(range(9, 18), voucher.items.all()):
        line_total = item.quantity * item.unit_cost
        total_amount += line_total

        ws[f"A{row}"] = item.description
        ws[f"L{row}"] = item.description
        ws[f"I{row}"] = float(line_total)
        ws[f"T{row}"] = float(line_total)

    amount = float(total_amount or voucher.amount_liquidated or voucher.amount_requested or 0)

    ws["I18"] = amount
    ws["T18"] = amount

    # ================= PURPOSE =================
    ws["C19"] = voucher.purpose
    ws["N19"] = voucher.purpose

    # ================= SIGNATORIES =================
    ws["C23"] = fullname
    ws["N23"] = fullname

    if administrator:
        admin_name = administrator.get_full_name().upper()
        ws["H23"] = admin_name
        ws["S23"] = admin_name

    # ================= DATE =================
    if voucher.purchase_date:
        date_str = voucher.purchase_date.strftime("%B %d, %Y")

        ws["D24"] = date_str
        ws["I24"] = date_str
        ws["O24"] = date_str
        ws["T24"] = date_str

    return wb
